"""
Dynamic multi-objective job shop scheduling - experiment analysis and management.

Core components:
1. MultiObjectiveManager: experiment data generation and management
2. ExperimentAnalysisSystem: analysis, visualization, and report generation

Dynamic metrics: always include hypervolume (HV) and coverage; runtime optional.
"""

import os
import sys
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook, Workbook
from pathlib import Path
from scipy import stats
from scipy.stats import friedmanchisquare, rankdata, wilcoxon, mannwhitneyu
import scikit_posthocs as sp
from typing import Dict, List, Tuple, Optional, Union
import json
from math import pi
import warnings
import logging
warnings.filterwarnings('ignore')

# Set fonts for CJK display
plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans', 'Arial Unicode MS', 'Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False
plt.rcParams['figure.dpi'] = 100
plt.rcParams['savefig.dpi'] = 300

# Configure structured logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('experiment_analysis.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


# ============================================================================
# Part 1: multi-objective experiment data generation (MultiObjectiveManager)
# ============================================================================

class MultiObjectiveManager:
    """Multi-objective manager: HV/coverage, optional runtime."""
    def __init__(self):
        self.all_experiment_data = {}           # All experiment data
        self.global_ideal = None                 # Global ideal point
        self.global_nadir = None                 # Global nadir point
        self.normalization_ranges = None         # Normalization ranges
        # Non-dominated sets
        self.algorithm_pareto_solutions = {}     # Per-algorithm non-dominated
        self.global_pareto_front = None          # Global Pareto front
    
    def add_experiment_data(self, rule_name, run_id, job_objectives):
        """Add experiment data."""
        if rule_name not in self.all_experiment_data:
            self.all_experiment_data[rule_name] = {}
        self.all_experiment_data[rule_name][run_id] = job_objectives
        # Clear caches
        self.algorithm_pareto_solutions = {}
        self.global_pareto_front = None
    
    def calculate_global_reference_points(self):
        """Compute global reference points for normalization."""
        all_solutions = []
        for rule_data in self.all_experiment_data.values():
            for run_data in rule_data.values():
                for job_data in run_data:
                    if 'objectives' in job_data:
                        all_solutions.append(job_data['objectives'])
        
        if all_solutions:
            all_solutions_array = np.array(all_solutions)
            self.global_ideal = np.min(all_solutions_array, axis=0)
            self.global_nadir = np.max(all_solutions_array, axis=0)
            self.normalization_ranges = self.global_nadir - self.global_ideal
            self.normalization_ranges[self.normalization_ranges == 0] = 1.0
        else:
            self.global_ideal = np.array([0, 0, 0])
            self.global_nadir = np.array([1, 1, 1])
            self.normalization_ranges = np.array([1, 1, 1])
    
    def normalize_solutions(self, solutions):
        """Normalize solutions to [0, 1]."""
        if self.global_ideal is None:
            self.calculate_global_reference_points()
        
        solutions_array = np.array(solutions)
        normalized = (solutions_array - self.global_ideal) / self.normalization_ranges
        return np.clip(normalized, 0, 1)
    
    def _calculate_pareto_mask(self, solutions):
        """Compute Pareto dominance mask (minimization)."""
        n = len(solutions)
        if n == 0:
            return np.array([], dtype=bool)
        
        mask = np.ones(n, dtype=bool)
        for i in range(n):
            if not mask[i]:
                continue
            for j in range(n):
                if i != j and mask[j]:
                    if np.all(solutions[j] <= solutions[i]) and np.any(solutions[j] < solutions[i]):
                        mask[i] = False
                        break
        return mask
    
    def calculate_pareto_fronts(self):
        """Key: compute per-algorithm and global Pareto fronts."""
        self.calculate_global_reference_points()
        
        # Initialize
        self.algorithm_pareto_solutions = {}
        all_solutions_with_info = []
        
        # Compute per-algorithm non-dominated sets
        for rule_name, rule_data in self.all_experiment_data.items():
            algorithm_solutions = []
            algorithm_info = []
            
            for run_id, run_data in rule_data.items():
                for job_data in run_data:
                    if 'objectives' in job_data:
                        algorithm_solutions.append(job_data['objectives'])
                        algorithm_info.append({
                            'rule_name': rule_name,
                            'objectives': job_data['objectives']
                        })
            
            if algorithm_solutions:
                solutions_array = np.array(algorithm_solutions)
                pareto_mask = self._calculate_pareto_mask(solutions_array)
                
                self.algorithm_pareto_solutions[rule_name] = {
                    'solutions': solutions_array[pareto_mask],
                    'info': [info for i, info in enumerate(algorithm_info) if pareto_mask[i]]
                }
                
                # Collect non-dominated for global front
                all_solutions_with_info.extend([
                    (sol, info) for sol, info in zip(solutions_array[pareto_mask], 
                                                      [info for i, info in enumerate(algorithm_info) if pareto_mask[i]])
                ])
            else:
                self.algorithm_pareto_solutions[rule_name] = {
                    'solutions': np.array([]),
                    'info': []
                }
        
        # Compute global Pareto front
        if all_solutions_with_info:
            all_solutions = np.array([item[0] for item in all_solutions_with_info])
            all_info = [item[1] for item in all_solutions_with_info]
            
            global_pareto_mask = self._calculate_pareto_mask(all_solutions)
            
            self.global_pareto_front = {
                'solutions': all_solutions[global_pareto_mask],
                'info': [info for i, info in enumerate(all_info) if global_pareto_mask[i]]
            }
        else:
            self.global_pareto_front = {'solutions': np.array([]), 'info': []}
    
    def calculate_hypervolume(self, solutions):
        """Compute hypervolume."""
        if len(solutions) < 1:
            return 0
        
        # Normalize
        normalized = self.normalize_solutions(solutions)
        fixed_ref_point = np.ones(3) * 1.01
        
        try:
            from pymoo.indicators.hv import HV
            ind = HV(ref_point=fixed_ref_point)
            return ind(normalized)
        except ImportError:
            # Simplified fallback
            if len(normalized) == 0:
                return 0
            
            solutions_sorted = normalized[np.argsort(normalized[:, 0])]
            hv_value = 0.0
            
            for sol in solutions_sorted:
                volume = 1.0
                for i in range(len(sol)):
                    if fixed_ref_point[i] > sol[i]:
                        volume *= (fixed_ref_point[i] - sol[i])
                    else:
                        volume = 0
                        break
                if volume > hv_value:
                    hv_value = volume
            
            return hv_value
    
    def calculate_coverage(self):
        """
        Compute coverage.
        Definition: fraction of global Pareto solutions contributed by each algorithm.
        Coverage = (algo solutions in global front / total global front) * 100%
        """
        if self.global_pareto_front is None:
            self.calculate_pareto_fronts()
        
        coverage = {}
        
        if self.global_pareto_front and len(self.global_pareto_front['info']) > 0:
            total_global = len(self.global_pareto_front['info'])
            
            # Count contributions per algorithm
            for rule_name in self.algorithm_pareto_solutions.keys():
                count = sum(1 for info in self.global_pareto_front['info'] 
                           if info['rule_name'] == rule_name)
                coverage[rule_name] = (count / total_global) * 100
        else:
            for rule_name in self.algorithm_pareto_solutions.keys():
                coverage[rule_name] = 0.0
        
        return coverage
    
    def get_metrics(self):
        """Return hypervolume and coverage metrics.
        Returns: (hypervolumes, coverages)
        """
        if not self.algorithm_pareto_solutions:
            self.calculate_pareto_fronts()
        
        # Hypervolume
        hypervolumes = {}
        for rule_name, data in self.algorithm_pareto_solutions.items():
            hypervolumes[rule_name] = self.calculate_hypervolume(data['solutions'])
        
        # Coverage
        coverages = self.calculate_coverage()
        
        return hypervolumes, coverages
    
    def save_to_excel(self, scenario_id, cyc, benchmark, hypervolumes, coverages, run_time=None, cpath='ablation'):
        """
        Save hypervolume, coverage, and runtime to Excel (runtime in its own sheet).
        
        Args:
            scenario_id: scenario id
            cyc: run index
            benchmark: algorithm list
            hypervolumes: {algorithm: value}
            coverages: {algorithm: value}
            run_time: {algorithm: seconds} (optional)
            cpath: subdirectory under test_result
        """
        
        excel_path = f"{sys.path[0]}\\test_result\\{cpath}\\{scenario_id}_metrics.xlsx"
        
        # Ensure directory exists
        directory = os.path.dirname(excel_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)
        
        try:
            wb = load_workbook(excel_path)
        except FileNotFoundError:
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
        
        # ========== 1. HV and coverage sheets (no runtime column) ==========
        metrics = [
            {'name': 'Hypervolume (HV)', 'data': hypervolumes, 'format': '{:.6f}'},
            {'name': 'Coverage', 'data': coverages, 'format': '{:.2f}%'}
        ]
        
        for metric in metrics:
            sheet_name = metric['name']
            data_dict = metric['data']
            
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
                # Header without runtime column
                headers = ['Scenario', 'Run'] + benchmark
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
            
            # Row without runtime column
            new_row = [scenario_id, cyc + 1]
            
            for algo in benchmark:
                value = data_dict.get(algo, 0)
                new_row.append(value)
            
            next_row = ws.max_row + 1
            for col_idx, value in enumerate(new_row, 1):
                ws.cell(row=next_row, column=col_idx, value=value)
        
        # ========== 2. Runtime sheet (optional) ==========
        if run_time is not None:
            sheet_name = "Runtime"
            # Check if runtime sheet exists
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
                # Header matches other sheets
                headers = ['Scenario', 'Run'] + benchmark
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
            
            # Runtime row
            new_row = [scenario_id, cyc + 1]
            for algo in benchmark:
                # Runtime value or 0
                value = run_time.get(algo, 0)
                new_row.append(value)
            
            # Write to next row
            next_row = ws.max_row + 1
            for col_idx, value in enumerate(new_row, 1):
                ws.cell(row=next_row, column=col_idx, value=value)
        
        # Save and close
        wb.save(excel_path)
        wb.close()


# ============================================================================
# Part 2: experiment analysis and visualization (ExperimentAnalysisSystem)
# ============================================================================

class ExperimentDataLoader:
    """Experiment data loader with optional runtime detection."""
    
    # Base metric config (always present)
    BASE_METRIC_CONFIG = {
        'Hypervolume (HV)': {
            'metric_key': 'hypervolume',
            'short_name': 'HV',
            'direction': 'higher',
            'sheet_name': 'Hypervolume (HV)',
            'description': 'Hypervolume (higher is better)'
        },
        'Coverage': {
            'metric_key': 'coverage',
            'short_name': 'Coverage',
            'direction': 'higher',
            'sheet_name': 'Coverage',
            'description': 'Coverage (higher is better)'
        }
    }
    
    # Optional metric config (runtime)
    OPTIONAL_METRIC_CONFIG = {
        'Runtime': {
            'metric_key': 'runtime',
            'short_name': 'Runtime',
            'direction': 'lower',  # Lower is better
            'sheet_name': 'Runtime',
            'description': 'Runtime (lower is better)'
        }
    }
    
    def __init__(self, data_path: str = 'test_result'):
        """
        Initialize data loader.
        
        Args:
            data_path: experiment data root (default: test_result)
        """
        # Project root
        if len(sys.path) > 0:
            project_root = sys.path[0]
        else:
            project_root = os.getcwd()
        
        self.data_path = Path(os.path.join(project_root, data_path))
        self.raw_data = {}
        self.processed_data = None
        self.summary_stats = {}
        self.benchmark_algorithms = None
        
        # Dynamic metric config
        self.METRIC_CONFIG = self.BASE_METRIC_CONFIG.copy()
        self.has_runtime = False  # Runtime present
        
    def get_metric_config(self):
        """Return metric config."""
        return self.METRIC_CONFIG.copy()
    
    def get_metric_names(self):
        """Return all metric names."""
        return [config['short_name'] for config in self.METRIC_CONFIG.values()]
    
    def get_metric_direction(self, metric_name: str) -> str:
        """Get metric direction."""
        for config in self.METRIC_CONFIG.values():
            if config['short_name'] == metric_name:
                return config['direction']
        return 'higher'  # Default
    
    def has_runtime_metric(self) -> bool:
        """Check if runtime is available."""
        return self.has_runtime
    
    def load_experiment_file(self, file_path: Union[str, Path]) -> Dict:
        """Load all sheets from one experiment file (runtime optional).
        Args:
            file_path: Excel file path
        Returns:
            Dict with all metric data
        """
        file_path = Path(file_path)
        scenario_id = file_path.stem.replace('_metrics', '')
        
        result = {
            'scenario': scenario_id,
            'file_path': file_path,
            'metrics': {}
        }
        
        try:
            wb = load_workbook(file_path, data_only=True)
            
            # Check runtime sheet
            if 'Runtime' in wb.sheetnames and not self.has_runtime:
                self.has_runtime = True
                # Add runtime to metric config
                self.METRIC_CONFIG['Runtime'] = self.OPTIONAL_METRIC_CONFIG['Runtime']
                logger.info("Runtime data detected; added to metrics")
            
            # Load all configured metrics
            for sheet_name, config in self.METRIC_CONFIG.items():
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    # Headers (algorithm names)
                    headers = []
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=1, column=col).value
                        if cell_value:
                            headers.append(cell_value)
                    
                    # Record benchmark algorithms
                    if self.benchmark_algorithms is None and len(headers) > 2:
                        # Skip first two columns
                        self.benchmark_algorithms = headers[2:]
                    
                    # Read data rows
                    data_rows = []
                    for row in range(2, ws.max_row + 1):
                        row_data = {}
                        row_has_data = False
                        
                        for col, header in enumerate(headers, 1):
                            cell_value = ws.cell(row=row, column=col).value
                            if cell_value is not None:
                                row_data[header] = cell_value
                                if col > 2:  # Skip Scenario/Run
                                    row_has_data = True
                        
                        if row_has_data:  # Only non-empty rows
                            data_rows.append(row_data)
                    
                    # Convert to DataFrame
                    if data_rows:
                        df = pd.DataFrame(data_rows)
                        # Ensure numeric columns
                        for col in df.columns:
                            if col not in ['Scenario', 'Run']:
                                df[col] = pd.to_numeric(df[col], errors='coerce')
                        
                        result['metrics'][config['short_name']] = {
                            'data': df,
                            'direction': config['direction'],
                            'sheet_name': sheet_name,
                            'description': config['description']
                        }
            
            wb.close()
            
        except Exception as e:
            print(f"  Error loading {file_path.name}: {e}")
            
        return result
    
    def load_all_experiments(self, sub_dir: str = 'ablation') -> Dict:
        """Load all experiment files under a directory.
        Args:
            sub_dir: subdirectory (default: ablation)
        Returns:
            Dict of all experiment data
        """
        target_dir = self.data_path / sub_dir
        
        if not target_dir.exists():
            print(f"Warning: directory not found: {target_dir}")
            # Fall back to current directory
            target_dir = self.data_path
        
        excel_files = list(target_dir.glob("*_metrics.xlsx"))
        
        if not excel_files:
            print(f"Error: no metrics files found in {target_dir}")
            return {}
        
        print(f"Found {len(excel_files)} experiment files")
        
        # Reset runtime detection
        self.has_runtime = False
        self.METRIC_CONFIG = self.BASE_METRIC_CONFIG.copy()
        
        for file_path in excel_files:
            print(f"Loading file: {file_path.name}")
            experiment_data = self.load_experiment_file(file_path)
            if experiment_data['metrics']:  # Only add files with data
                scenario = experiment_data['scenario']
                self.raw_data[scenario] = experiment_data
            else:
                print(f"  Warning: {file_path.name} has no valid data")
        
        if self.raw_data:
            self._process_data()
            metric_count = len(self.get_metric_names())
            runtime_status = "with runtime" if self.has_runtime else "no runtime"
            print(f"Loaded {len(self.raw_data)} scenarios with {metric_count} metrics ({runtime_status})")
        else:
            print("Error: no data loaded")
        
        return self.raw_data
    
    def _process_data(self):
        """Process raw data into analysis-ready format."""
        
        all_data = []
        
        for scenario, scenario_data in self.raw_data.items():
            for metric_name, metric_info in scenario_data['metrics'].items():
                df = metric_info['data'].copy()
                
                # Long format
                for algo in self.benchmark_algorithms:
                    if algo in df.columns:
                        values = df[algo].dropna().values
                        for i, value in enumerate(values):
                            all_data.append({
                                'Scenario': scenario,
                                'Algorithm': algo,
                                'Metric': metric_name,
                                'Value': value,
                                'Run': i + 1,
                                'Direction': metric_info['direction']
                            })
        
        if all_data:
            self.processed_data = pd.DataFrame(all_data)
            self._calculate_summary_stats()
            print(f"Processing complete: {len(self.processed_data)} records")
        else:
            print("Warning: no valid data after processing")
    
    def _calculate_summary_stats(self):
        """Compute summary statistics."""
        if self.processed_data is None or len(self.processed_data) == 0:
            return
        
        # Mean by scenario/algorithm/metric
        self.summary_stats['mean'] = self.processed_data.groupby(
            ['Scenario', 'Algorithm', 'Metric']
        )['Value'].mean().reset_index()
        
        # Standard deviation
        self.summary_stats['std'] = self.processed_data.groupby(
            ['Scenario', 'Algorithm', 'Metric']
        )['Value'].std().reset_index()
        
        # Overall mean
        self.summary_stats['overall_mean'] = self.processed_data.groupby(
            ['Algorithm', 'Metric']
        )['Value'].mean().reset_index()
        
        # Overall std
        self.summary_stats['overall_std'] = self.processed_data.groupby(
            ['Algorithm', 'Metric']
        )['Value'].std().reset_index()
        
        # Median
        self.summary_stats['overall_median'] = self.processed_data.groupby(
            ['Algorithm', 'Metric']
        )['Value'].median().reset_index()
        
        # Min/max
        self.summary_stats['overall_min'] = self.processed_data.groupby(
            ['Algorithm', 'Metric']
        )['Value'].min().reset_index()
        
        self.summary_stats['overall_max'] = self.processed_data.groupby(
            ['Algorithm', 'Metric']
        )['Value'].max().reset_index()
    
    def get_data_for_metric(self, metric_name: str) -> pd.DataFrame:
        """Return all data for a metric.
        Args:
            metric_name: metric name (HV/Coverage/Runtime)
        Returns:
            DataFrame with metric data
        """
        if self.processed_data is None:
            return pd.DataFrame()
        
        return self.processed_data[self.processed_data['Metric'] == metric_name].copy()
    
    def get_algorithm_ranking(self) -> pd.DataFrame:
        """Compute overall algorithm ranking (all metrics)."""
        if self.summary_stats is None or 'overall_mean' not in self.summary_stats:
            return pd.DataFrame()
        
        overall_mean = self.summary_stats['overall_mean'].copy()
        
        # Normalize
        normalized_values = []
        
        for metric in overall_mean['Metric'].unique():
            metric_df = overall_mean[overall_mean['Metric'] == metric].copy()
            direction = self.get_metric_direction(metric)
            
            min_val = metric_df['Value'].min()
            max_val = metric_df['Value'].max()
            
            if max_val == min_val:
                metric_df['Normalized'] = 0.5
            elif direction == 'higher':
                metric_df['Normalized'] = (metric_df['Value'] - min_val) / (max_val - min_val)
            else:  # lower
                metric_df['Normalized'] = 1 - (metric_df['Value'] - min_val) / (max_val - min_val)
            
            normalized_values.append(metric_df)
        
        if normalized_values:
            normalized_df = pd.concat(normalized_values, ignore_index=True)
            
            # Composite score (equal weights)
            ranking = normalized_df.groupby('Algorithm')['Normalized'].mean().reset_index()
            ranking = ranking.sort_values('Normalized', ascending=False)
            ranking['Rank'] = range(1, len(ranking) + 1)
            ranking['Normalized'] = ranking['Normalized'].round(4)
            
            return ranking
        
        return pd.DataFrame()
    
    def get_algorithm_stats(self) -> pd.DataFrame:
        """Return detailed algorithm statistics."""
        if self.summary_stats is None:
            return pd.DataFrame()
        
        stats_list = []
        
        for algo in self.benchmark_algorithms:
            for metric in self.get_metric_names():
                mean_df = self.summary_stats['overall_mean']
                std_df = self.summary_stats['overall_std']
                median_df = self.summary_stats['overall_median']
                min_df = self.summary_stats['overall_min']
                max_df = self.summary_stats['overall_max']
                
                mean_val = mean_df[(mean_df['Algorithm'] == algo) & (mean_df['Metric'] == metric)]['Value'].values
                std_val = std_df[(std_df['Algorithm'] == algo) & (std_df['Metric'] == metric)]['Value'].values
                median_val = median_df[(median_df['Algorithm'] == algo) & (median_df['Metric'] == metric)]['Value'].values
                min_val = min_df[(min_df['Algorithm'] == algo) & (min_df['Metric'] == metric)]['Value'].values
                max_val = max_df[(max_df['Algorithm'] == algo) & (max_df['Metric'] == metric)]['Value'].values
                
                if len(mean_val) > 0:
                    stats_list.append({
                        'Algorithm': algo,
                        'Metric': metric,
                        'Mean': mean_val[0],
                        'Std': std_val[0] if len(std_val) > 0 else np.nan,
                        'Median': median_val[0] if len(median_val) > 0 else np.nan,
                        'Min': min_val[0] if len(min_val) > 0 else np.nan,
                        'Max': max_val[0] if len(max_val) > 0 else np.nan
                    })
        
        return pd.DataFrame(stats_list)
    
    def generate_detailed_scenario_metrics(self, output_path: Union[str, Path], 
                                            significance_levels: Dict = None):
        """Generate per-scenario metric Excel (all available metrics).
        Args:
            output_path: output Excel path
            significance_levels: significance levels
        """
        if self.processed_data is None or len(self.processed_data) == 0:
            logger.error("No valid data; cannot generate report")
            return
        
        # Default significance levels
        if significance_levels is None:
            significance_levels = {'**': 0.01, '*': 0.05, '†': 0.10}
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Scenarios and algorithms
        scenarios = sorted(self.processed_data['Scenario'].unique())
        algorithms = self.benchmark_algorithms if self.benchmark_algorithms else \
                sorted(self.processed_data['Algorithm'].unique())
        
        # Build metric info
        metrics_info = {}
        
        # HV
        if 'HV' in self.processed_data['Metric'].unique():
            metrics_info['HV'] = {
                'sheet_name': 'Hypervolume (HV)', 
                'direction': 'higher',
                'description': 'Hypervolume (higher is better)'
            }
        
        # Coverage
        if 'Coverage' in self.processed_data['Metric'].unique():
            metrics_info['Coverage'] = {
                'sheet_name': 'Coverage', 
                'direction': 'higher',
                'description': 'Coverage (higher is better)'
            }
        
        # Runtime
        if self.has_runtime and 'Runtime' in self.processed_data['Metric'].unique():
            metrics_info['Runtime'] = {
                'sheet_name': 'Runtime', 
                'direction': 'lower',  # Lower is better
                'description': 'Runtime (lower is better)'
            }
        
        # Track sheet creation
        sheet_created = False
        
        # Store sheet data
        all_sheets = {}
        
        for metric_name, metric_cfg in metrics_info.items():
            metric_data = self.processed_data[self.processed_data['Metric'] == metric_name]
            
            if metric_data.empty:
                logger.warning(f"No data for metric {metric_name}; skipping")
                continue
            
            logger.info(f"Processing metric: {metric_name} ({metric_cfg['description']})")
            
            # Pivot: rows=scenarios, cols=algorithms, values=mean
            pivot_mean = metric_data.groupby(['Scenario', 'Algorithm'])['Value'].mean().unstack()
            pivot_std = metric_data.groupby(['Scenario', 'Algorithm'])['Value'].std().unstack()
            
            # Ensure all algorithms exist
            for algo in algorithms:
                if algo not in pivot_mean.columns:
                    pivot_mean[algo] = np.nan
                    pivot_std[algo] = np.nan
            
            # Reorder columns
            pivot_mean = pivot_mean[algorithms]
            pivot_std = pivot_std[algorithms]
            
            # Display strings and best flags
            display_df = pd.DataFrame(index=pivot_mean.index, columns=pivot_mean.columns)
            best_flags = pd.DataFrame(index=pivot_mean.index, columns=pivot_mean.columns, dtype=bool)
            
            from scipy import stats
            
            # Format per scenario
            for scenario in pivot_mean.index:
                # Mean per algorithm (exclude NaN)
                scenario_means = pivot_mean.loc[scenario].copy()
                scenario_means = scenario_means[~pd.isna(scenario_means)]
                
                if len(scenario_means) == 0:
                    # No data for this scenario
                    for algo in algorithms:
                        display_df.loc[scenario, algo] = 'N/A'
                        best_flags.loc[scenario, algo] = False
                    continue
                
                # Best by direction
                if metric_cfg['direction'] == 'higher':
                    best_value = scenario_means.max()
                    best_algo = scenario_means.idxmax()
                else:  # lower
                    best_value = scenario_means.min()
                    best_algo = scenario_means.idxmin()
                
                logger.debug(f"Scenario {scenario}: {metric_name} best={best_algo}, value={best_value:.4f}")
                
                # Format each algorithm
                for algo in algorithms:
                    mean_val = pivot_mean.loc[scenario, algo]
                    std_val = pivot_std.loc[scenario, algo]
                    
                    if pd.isna(mean_val) or pd.isna(std_val):
                        display_df.loc[scenario, algo] = 'N/A'
                        best_flags.loc[scenario, algo] = False
                        continue
                    
                    # Base string: mean ± std (4 decimals)
                    base_str = f"{mean_val:.4f} ± {std_val:.4f}"
                    
                    # Mark best
                    is_best = (algo == best_algo)
                    best_flags.loc[scenario, algo] = is_best
                    
                    # Significance superscript vs best
                    superscript = ''
                    if not is_best and not pd.isna(mean_val) and not pd.isna(best_value):
                        try:
                            # Raw values for both algorithms
                            algo_data = metric_data[
                                (metric_data['Scenario'] == scenario) & 
                                (metric_data['Algorithm'] == algo)
                            ]['Value'].values
                            
                            best_data = metric_data[
                                (metric_data['Scenario'] == scenario) & 
                                (metric_data['Algorithm'] == best_algo)
                            ]['Value'].values
                            
                            if len(algo_data) > 1 and len(best_data) > 1:
                                # t-test
                                _, p_value = stats.ttest_ind(algo_data, best_data, equal_var=False)
                                
                                # Add superscript by p-value
                                for sym, level in sorted(significance_levels.items(), 
                                                        key=lambda x: x[1]):
                                    if p_value < level:
                                        superscript = sym
                                        break
                                
                                if superscript:
                                    logger.debug(f"  {algo} vs best {best_algo}: p={p_value:.4f} {superscript}")
                        except Exception as e:
                            logger.debug(f"Significance test failed: {e}")
                    
                    # Final display string
                    display_df.loc[scenario, algo] = base_str + superscript
            
            # Store sheet data and flags
            all_sheets[metric_cfg['sheet_name']] = {
                'data': display_df,
                'best_flags': best_flags,
                'direction': metric_cfg['direction'],
                'description': metric_cfg['description']
            }
            sheet_created = True
        
        # Default sheet when none created
        if not sheet_created:
            logger.warning("No valid metric data; creating default sheet")
            default_df = pd.DataFrame(['No valid data'], columns=['Notice'])
            all_sheets['Summary'] = {
                'data': default_df,
                'best_flags': pd.DataFrame(),
                'direction': 'higher',
                'description': 'No data'
            }
        
        # ==================== Write Excel ====================
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                from openpyxl.styles import Font, Alignment
                
                # Write all sheets
                for sheet_name, sheet_content in all_sheets.items():
                    display_df = sheet_content['data']
                    best_flags = sheet_content['best_flags']
                    description = sheet_content.get('description', '')
                    
                    # Write data
                    display_df.to_excel(writer, sheet_name=sheet_name)
                    
                    # Styles
                    workbook = writer.book
                    worksheet = writer.sheets[sheet_name]
                    
                    # Header note with direction
                    if description:
                        worksheet.cell(row=1, column=len(display_df.columns) + 2, value=description)
                        worksheet.cell(row=1, column=len(display_df.columns) + 2).font = Font(italic=True)
                    
                    # Column widths
                    worksheet.column_dimensions['A'].width = 20
                    from openpyxl.utils import get_column_letter
                    for col_idx in range(2, len(display_df.columns) + 2):
                        col_letter = get_column_letter(col_idx)
                        worksheet.column_dimensions[col_letter].width = 22
                    
                    # Alignment
                    center_align = Alignment(horizontal='center', vertical='center')
                    bold_font = Font(bold=True)
                    
                    # Header style
                    for col_idx, algo in enumerate(display_df.columns, start=2):
                        cell = worksheet.cell(row=1, column=col_idx)
                        cell.font = bold_font
                        cell.alignment = center_align
                        cell.value = algo
                    
                    # Data cell style
                    for row_idx, scenario in enumerate(display_df.index, start=2):
                        # Scenario name cell
                        scene_cell = worksheet.cell(row=row_idx, column=1)
                        scene_cell.value = str(scenario)
                        scene_cell.alignment = center_align
                        
                        # Data cells
                        for col_idx, algo in enumerate(display_df.columns, start=2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = center_align
                            
                            # Bold best values
                            if not best_flags.empty and scenario in best_flags.index and algo in best_flags.columns:
                                if best_flags.loc[scenario, algo]:
                                    cell.font = bold_font
                
                # ==================== Significance notes ====================
                sign_data = [
                    ['Symbol', 'Threshold', 'Description'],
                    ['**', f'p < {significance_levels["**"]}', 'Highly significant'],
                    ['*', f'p < {significance_levels["*"]}', 'Significant'],
                    ['†', f'p < {significance_levels["†"]}', 'Marginal'],
                    ['Bold', '-', 'Best algorithm in scenario'],
                    ['', '', ''],
                    ['Metric direction:', '', '']
                ]
                
                # Add metric direction notes
                for metric_name, metric_cfg in metrics_info.items():
                    sign_data.append([metric_cfg['sheet_name'], 
                                     'Higher is better' if metric_cfg['direction'] == 'higher' else 'Lower is better', 
                                     ''])
                
                sign_df = pd.DataFrame(sign_data[1:], columns=sign_data[0])
                sign_df.to_excel(writer, sheet_name='Significance', index=False)
                
                # Significance sheet style
                worksheet = writer.sheets['Significance']
                worksheet.column_dimensions['A'].width = 15
                worksheet.column_dimensions['B'].width = 20
                worksheet.column_dimensions['C'].width = 25
                
                # Bold header
                for col_idx in range(1, 4):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.font = Font(bold=True)
        
        except Exception as e:
            logger.error(f"Failed to write Excel: {e}")
            # Fallback: CSV
            csv_path = output_path.with_suffix('.csv')
            logger.info(f"Trying CSV: {csv_path}")
            
            for sheet_name, sheet_content in all_sheets.items():
                display_df = sheet_content['data']
                csv_sheet_path = csv_path.parent / f"{csv_path.stem}_{sheet_name}.csv"
                display_df.to_csv(csv_sheet_path, encoding='utf-8-sig')
                logger.info(f"Saved: {csv_sheet_path}")
            
            return str(csv_path)
        
        logger.info(f"Scenario metric report generated: {output_path}")
        return str(output_path)
    
    def generate_win_rate_statistics(self) -> pd.DataFrame:
        """
        Generate win-rate statistics (all available metrics).
        
        Returns:
            DataFrame with:
                - Algorithm
                - BestCount: scenarios ranked first
                - BestRate
                - SignificantlyBetterCount (p<0.05)
                - SignificantlyBetterRate
                - SignificantlyWorseCount
                - SignificantlyWorseRate
                - OverallWinRate
        """
        if self.processed_data is None or len(self.processed_data) == 0:
            logger.error("No valid data; cannot generate win-rate stats")
            return pd.DataFrame()
        
        metrics = self.get_metric_names()
        
        scenarios = self.processed_data['Scenario'].unique()
        algorithms = self.benchmark_algorithms if self.benchmark_algorithms else \
                self.processed_data['Algorithm'].unique()
        
        # Init stats
        stats = {algo: {
            'BestCount': 0,
            'SignificantlyBetterCount': 0,
            'SignificantlyWorseCount': 0,
            'ScenarioCount': 0
        } for algo in algorithms}
        
        from scipy import stats as scipy_stats
        
        # ==================== Loop metrics/scenarios ====================
        for metric in metrics:
            metric_data = self.processed_data[self.processed_data['Metric'] == metric]
            direction = self.get_metric_direction(metric)
            
            for scenario in scenarios:
                scenario_data = metric_data[metric_data['Scenario'] == scenario]
                
                if scenario_data.empty:
                    continue
                
                # Mean per algorithm in this scenario
                algo_means = {}
                algo_values = {}
                
                for algo in algorithms:
                    algo_data = scenario_data[scenario_data['Algorithm'] == algo]['Value'].values
                    if len(algo_data) > 0:
                        algo_means[algo] = np.mean(algo_data)
                        algo_values[algo] = algo_data
                        stats[algo]['ScenarioCount'] += 1
                
                if not algo_means:
                    continue
                
                # Best by direction
                if direction == 'higher':
                    best_algo = max(algo_means, key=algo_means.get)
                else:  # lower
                    best_algo = min(algo_means, key=algo_means.get)
                
                best_value = algo_means[best_algo]
                
                # Update best count
                stats[best_algo]['BestCount'] += 1
                
                # Significance tests
                for algo in algorithms:
                    if algo == best_algo or algo not in algo_values:
                        continue
                    
                    algo_data = algo_values[algo]
                    best_data = algo_values[best_algo]
                    
                    if len(algo_data) >= 2 and len(best_data) >= 2:
                        try:
                            _, p_value = scipy_stats.ttest_ind(algo_data, best_data, equal_var=False)
                            
                            if p_value < 0.05:
                                # Significantly worse than best
                                stats[algo]['SignificantlyWorseCount'] += 1
                                stats[best_algo]['SignificantlyBetterCount'] += 1
                        except:
                            pass
        
        # ==================== Build table ====================
        total_scenarios = len(scenarios) * len(metrics)
        
        result = []
        for algo in algorithms:
            s = stats[algo]
            
            # Rates
            best_ratio = s['BestCount'] / total_scenarios if total_scenarios > 0 else 0
            sig_better_ratio = s['SignificantlyBetterCount'] / total_scenarios if total_scenarios > 0 else 0
            sig_worse_ratio = s['SignificantlyWorseCount'] / total_scenarios if total_scenarios > 0 else 0
            
            # Overall win rate
            comprehensive_win_rate = (s['BestCount'] + 0.5 * s['SignificantlyBetterCount']) / total_scenarios if total_scenarios > 0 else 0
            
            result.append({
                'Algorithm': algo,
                'BestCount': s['BestCount'],
                'BestRate': f"{best_ratio:.2%}",
                'SignificantlyBetterCount': s['SignificantlyBetterCount'],
                'SignificantlyBetterRate': f"{sig_better_ratio:.2%}",
                'SignificantlyWorseCount': s['SignificantlyWorseCount'],
                'SignificantlyWorseRate': f"{sig_worse_ratio:.2%}",
                'OverallWinRate': f"{comprehensive_win_rate:.2%}",
                'OverallScore': comprehensive_win_rate  # for sorting
            })
        
        # Sort by overall score
        result_df = pd.DataFrame(result)
        result_df = result_df.sort_values('OverallScore', ascending=False)
        result_df = result_df.drop('OverallScore', axis=1)
        result_df['Rank'] = range(1, len(result_df) + 1)
        
        # Reorder columns
        column_order = ['Rank', 'Algorithm', 'BestCount', 'BestRate',
                'SignificantlyBetterCount', 'SignificantlyBetterRate',
                'SignificantlyWorseCount', 'SignificantlyWorseRate', 'OverallWinRate']
        result_df = result_df[column_order]
        
        logger.info("Win-rate statistics generated")
        return result_df

    def get_algorithm_ranking_with_stability(self, stability_weight: float = 0.3) -> pd.DataFrame:
        """Overall ranking with stability weighting.
        Args:
            stability_weight: weight in [0,1]
        Returns:
            DataFrame with Algorithm/PerformanceScore/StabilityScore/OverallScore/Rank
        """
        if self.summary_stats is None or 'overall_mean' not in self.summary_stats:
            logger.error("No summary stats; load data first")
            return pd.DataFrame()
        
        overall_mean = self.summary_stats['overall_mean'].copy()
        overall_std = self.summary_stats['overall_std'].copy()
        
        metrics = self.get_metric_names()
        algorithms = self.benchmark_algorithms if self.benchmark_algorithms else \
                overall_mean['Algorithm'].unique()
        
        result_data = []
        
        for algo in algorithms:
            algo_scores = {
                'Algorithm': algo,
                'PerformanceScore': 0,
                'StabilityScore': 0,
                'OverallScore': 0
            }
            
            valid_metrics = 0
            
            for metric in metrics:
                # Mean/std for this metric
                mean_val = overall_mean[
                    (overall_mean['Algorithm'] == algo) & 
                    (overall_mean['Metric'] == metric)
                ]['Value'].values
                
                std_val = overall_std[
                    (overall_std['Algorithm'] == algo) & 
                    (overall_std['Metric'] == metric)
                ]['Value'].values
                
                if len(mean_val) == 0 or len(std_val) == 0:
                    continue
                
                mean_val = mean_val[0]
                std_val = std_val[0]
                direction = self.get_metric_direction(metric)
                
                # === 1. Performance score (normalized mean) ===
                all_means = overall_mean[overall_mean['Metric'] == metric]['Value'].values
                min_mean, max_mean = np.min(all_means), np.max(all_means)
                
                if max_mean == min_mean:
                    perf_score = 0.5
                elif direction == 'higher':
                    perf_score = (mean_val - min_mean) / (max_mean - min_mean)
                else:  # lower
                    perf_score = 1 - (mean_val - min_mean) / (max_mean - min_mean)
                
                # === 2. Stability score via CV (lower is better) ===
                if mean_val != 0 and not np.isnan(std_val):
                    cv = std_val / abs(mean_val)
                else:
                    cv = 1.0
                
                # CV range across algorithms
                all_cvs = []
                for other_algo in algorithms:
                    other_mean = overall_mean[
                        (overall_mean['Algorithm'] == other_algo) & 
                        (overall_mean['Metric'] == metric)
                    ]['Value'].values
                    other_std = overall_std[
                        (overall_std['Algorithm'] == other_algo) & 
                        (overall_std['Metric'] == metric)
                    ]['Value'].values
                    
                    if len(other_mean) > 0 and len(other_std) > 0 and other_mean[0] != 0:
                        all_cvs.append(other_std[0] / abs(other_mean[0]))
                    else:
                        all_cvs.append(1.0)
                
                min_cv, max_cv = np.min(all_cvs), np.max(all_cvs)
                
                if max_cv == min_cv:
                    stability_score = 0.5
                else:
                    # Lower CV -> higher score
                    stability_score = 1 - (cv - min_cv) / (max_cv - min_cv)
                
                # Accumulate
                algo_scores['PerformanceScore'] += perf_score
                algo_scores['StabilityScore'] += stability_score
                valid_metrics += 1
            
            if valid_metrics > 0:
                algo_scores['PerformanceScore'] /= valid_metrics
                algo_scores['StabilityScore'] /= valid_metrics
                # Weighted overall score
                algo_scores['OverallScore'] = (1 - stability_weight) * algo_scores['PerformanceScore'] + \
                                        stability_weight * algo_scores['StabilityScore']
                result_data.append(algo_scores)
        
        # Build DataFrame and sort
        result_df = pd.DataFrame(result_data)
        if not result_df.empty:
            result_df = result_df.sort_values('OverallScore', ascending=False)
            result_df['Rank'] = range(1, len(result_df) + 1)
            result_df['PerformanceScore'] = result_df['PerformanceScore'].round(4)
            result_df['StabilityScore'] = result_df['StabilityScore'].round(4)
            result_df['OverallScore'] = result_df['OverallScore'].round(4)
            
            # Reorder columns
            column_order = ['Rank', 'Algorithm', 'OverallScore', 'PerformanceScore', 'StabilityScore']
            result_df = result_df[column_order]
        
        logger.info(f"Stability-weighted ranking generated (weight={stability_weight})")
        return result_df


class ExperimentAnalyzer:
    """Experiment analyzer (HV/Coverage/Runtime)."""
    
    def __init__(self, data_loader: ExperimentDataLoader):
        """
        Initialize analyzer.
        Args:
            data_loader: ExperimentDataLoader
        """
        self.data_loader = data_loader
        self.processed_data = data_loader.processed_data
        self.summary_stats = data_loader.summary_stats
        self.benchmark_algorithms = data_loader.benchmark_algorithms
        
    def friedman_test(self, metric_name: str) -> Dict:
        """
        Friedman test for related samples.
        Args:
            metric_name: metric name
        Returns:
            Dict with test results
        """
        df = self.data_loader.get_data_for_metric(metric_name)
        
        if df.empty:
            return {'error': 'No data', 'metric': metric_name}
        
        # Scenario x algorithm matrix
        pivot_df = df.groupby(['Scenario', 'Algorithm'])['Value'].mean().reset_index()
        pivot_df = pivot_df.pivot(index='Scenario', columns='Algorithm', values='Value')
        
        # Drop NaN columns
        pivot_df = pivot_df.dropna(axis=1)
        
        if pivot_df.shape[1] < 2:
            return {'error': 'Not enough algorithms', 'metric': metric_name}
        
        if pivot_df.shape[0] < 2:
            return {'error': 'Not enough scenarios', 'metric': metric_name}
        
        algorithms = pivot_df.columns.tolist()
        data_matrix = [pivot_df[algo].values for algo in algorithms]
        
        try:
            stat, p_value = friedmanchisquare(*data_matrix)
            
            result = {
                'metric': metric_name,
                'statistic': stat,
                'p_value': p_value,
                'significant': p_value < 0.05,
                'algorithms': algorithms,
                'data_matrix': pivot_df
            }
            
            return result
        except Exception as e:
            return {'error': str(e), 'metric': metric_name}
        
    def nemenyi_posthoc(self, metric_name: str) -> pd.DataFrame:
        """
        Nemenyi post-hoc test.
        Args:
            metric_name: metric name
        Returns:
            DataFrame of pairwise p-values
        """
        friedman_result = self.friedman_test(metric_name)
        
        if 'error' in friedman_result:
            print(f"Nemenyi test: {metric_name} - {friedman_result.get('error', 'Unknown error')}")
            return pd.DataFrame()
        
        pivot_df = friedman_result['data_matrix'].copy()
        
        try:
            # 1) Drop any NaN rows/columns
            pivot_df = pivot_df.dropna(axis=1, how='any')
            pivot_df = pivot_df.dropna(axis=0, how='any')
            
            # 2) Require at least 2 algorithms
            if pivot_df.shape[1] < 2:
                print(f"Nemenyi test: {metric_name} insufficient algorithms ({pivot_df.shape[1]})")
                return pd.DataFrame()
            
            # 3) Valid algorithms
            algorithms = pivot_df.columns.tolist()
            n_algorithms = len(algorithms)
            
            print(f"Nemenyi test: {metric_name} - algorithms={n_algorithms}")
            
            # 4) scikit_posthoc expects (samples, treatments)
            data_for_test = pivot_df.values.T
            
            # 5) Run Nemenyi
            nemenyi_result = sp.posthoc_nemenyi_friedman(data_for_test)
            
            # 6) DataFrame with labels
            result_df = pd.DataFrame(
                nemenyi_result,
                index=algorithms[:nemenyi_result.shape[0]],
                columns=algorithms[:nemenyi_result.shape[1]]
            )
            
            return result_df
            
        except Exception as e:
            print(f"Nemenyi test error: {e}")
            
            # Fallback: manual computation
            try:
                print(f"Trying manual Nemenyi for {metric_name}...")
                
                from scipy.stats import rankdata, norm
                import numpy as np
                
                # Clean data
                pivot_df = pivot_df.dropna(axis=1, how='any')
                pivot_df = pivot_df.dropna(axis=0, how='any')
                
                if pivot_df.shape[1] < 2:
                    return pd.DataFrame()
                
                n_samples = pivot_df.shape[0]
                n_treatments = pivot_df.shape[1]
                algorithms = pivot_df.columns.tolist()
                
                # Rank per row
                ranks = np.zeros((n_samples, n_treatments))
                for i in range(n_samples):
                    ranks[i, :] = rankdata(pivot_df.iloc[i, :].values)
                
                # Mean ranks
                avg_ranks = np.mean(ranks, axis=0)
                
                # p-value matrix
                p_matrix = np.ones((n_treatments, n_treatments))
                for i in range(n_treatments):
                    for j in range(i + 1, n_treatments):
                        z = abs(avg_ranks[i] - avg_ranks[j]) / np.sqrt((n_treatments * (n_treatments + 1)) / (6 * n_samples))
                        p_value = 2 * (1 - norm.cdf(z))
                        p_matrix[i, j] = p_value
                        p_matrix[j, i] = p_value
                
                result_df = pd.DataFrame(
                    p_matrix,
                    index=algorithms,
                    columns=algorithms
                )
                
                print("Manual Nemenyi completed")
                return result_df
                
            except Exception as e2:
                print(f"Manual Nemenyi also failed: {e2}")
                return pd.DataFrame()
        
    def wilcoxon_pairwise(self, metric_name: str, algorithm1: str, algorithm2: str) -> Dict:
        """
        Wilcoxon signed-rank test for paired samples.
        Args:
            metric_name: metric name
            algorithm1: algorithm 1
            algorithm2: algorithm 2
        Returns:
            Dict with test results
        """
        df = self.data_loader.get_data_for_metric(metric_name)
        
        if df.empty:
            return {'error': 'No data'}
        
        # Paired data across scenarios
        algo1_data = df[df['Algorithm'] == algorithm1].groupby('Scenario')['Value'].mean()
        algo2_data = df[df['Algorithm'] == algorithm2].groupby('Scenario')['Value'].mean()
        
        # Align scenarios
        common_scenes = algo1_data.index.intersection(algo2_data.index)
        
        if len(common_scenes) < 2:
            return {'error': 'Insufficient paired samples'}
        
        x = algo1_data[common_scenes].values
        y = algo2_data[common_scenes].values
        
        try:
            stat, p_value = wilcoxon(x, y, zero_method='wilcox', correction=False)
            
            # Effect size
            n = len(common_scenes)
            effect_size = abs(stat) / (n * (n + 1) / 4)
            
            result = {
                'metric': metric_name,
                'algorithm1': algorithm1,
                'algorithm2': algorithm2,
                'statistic': stat,
                'p_value': p_value,
                'significant': p_value < 0.05,
                'effect_size': effect_size,
                'n_pairs': n,
                'mean1': np.mean(x),
                'mean2': np.mean(y),
                'diff': np.mean(x) - np.mean(y)
            }
            
            return result
        except Exception as e:
            return {'error': str(e)}
    
    def calculate_performance_profile(self, metric_name: str, tau_range: np.ndarray = None) -> pd.DataFrame:
        """
        Compute performance profile data.
        Args:
            metric_name: metric name
            tau_range: threshold range
        Returns:
            DataFrame with profile data
        """
        df = self.data_loader.get_data_for_metric(metric_name)
        
        if df.empty:
            return pd.DataFrame()
        
        direction = self.data_loader.get_metric_direction(metric_name)
        
        # Mean per scenario/algorithm
        algo_perf = df.groupby(['Scenario', 'Algorithm'])['Value'].mean().reset_index()
        
        # Best per scenario
        if direction == 'higher':
            best_perf = algo_perf.groupby('Scenario')['Value'].max().reset_index()
        else:  # lower
            best_perf = algo_perf.groupby('Scenario')['Value'].min().reset_index()
            
        best_perf.columns = ['Scenario', 'BestValue']
        algo_perf = algo_perf.merge(best_perf, on='Scenario')
        
        if direction == 'higher':
            algo_perf['Ratio'] = algo_perf['BestValue'] / algo_perf['Value']
        else:  # lower
            algo_perf['Ratio'] = algo_perf['Value'] / algo_perf['BestValue']
        
        # Threshold range
        if tau_range is None:
            max_ratio = min(3.0, algo_perf['Ratio'].max() * 1.1)
            tau_range = np.linspace(1.0, max_ratio, 50)
        
        # Profile per algorithm
        profile_data = []
        
        for algo in algo_perf['Algorithm'].unique():
            algo_ratios = algo_perf[algo_perf['Algorithm'] == algo]['Ratio'].values
            
            for tau in tau_range:
                prob = np.mean(algo_ratios <= tau)
                profile_data.append({
                    'Algorithm': algo,
                    'tau': tau,
                    'Probability': prob
                })
        
        return pd.DataFrame(profile_data)
    
    def calculate_statistical_summary(self) -> Dict:
        """
        Compute full statistical summary (all metrics).
        Returns:
            Dict with summary info
        """
        summary = {
            'friedman_tests': {},
            'ranking': self.data_loader.get_algorithm_ranking(),
            'best_algorithms': {},
            'pairwise_comparisons': {}
        }
        
        # Metric config
        metric_config = self.data_loader.get_metric_config()
        
        # Friedman test per metric
        for config in metric_config.values():
            metric_name = config['short_name']
            friedman_result = self.friedman_test(metric_name)
            summary['friedman_tests'][metric_name] = friedman_result
            
            if 'error' not in friedman_result and friedman_result['significant']:
                # Nemenyi post-hoc
                nemenyi_result = self.nemenyi_posthoc(metric_name)
                summary['friedman_tests'][metric_name]['nemenyi'] = nemenyi_result
            
            # Best algorithm per metric
            df = self.data_loader.get_data_for_metric(metric_name)
            if not df.empty:
                algo_means = df.groupby('Algorithm')['Value'].mean()
                
                if config['direction'] == 'higher':
                    best_algo = algo_means.idxmax()
                    best_value = algo_means.max()
                else:
                    best_algo = algo_means.idxmin()
                    best_value = algo_means.min()
                
                summary['best_algorithms'][metric_name] = {
                    'algorithm': best_algo,
                    'value': best_value,
                    'direction': config['direction']
                }
        
        # Pairwise comparisons for top-5
        if not summary['ranking'].empty:
            top5 = summary['ranking'].head(min(5, len(summary['ranking'])))['Algorithm'].tolist()
            
            for i, algo1 in enumerate(top5):
                for algo2 in top5[i+1:]:
                    key = f"{algo1} vs {algo2}"
                    summary['pairwise_comparisons'][key] = {}
                    
                    for config in metric_config.values():
                        metric_name = config['short_name']
                        wilcoxon_result = self.wilcoxon_pairwise(metric_name, algo1, algo2)
                        summary['pairwise_comparisons'][key][metric_name] = wilcoxon_result
        
        return summary


class ExperimentVisualizer:
    """Experiment visualizer (HV/Coverage/Runtime)."""
    
    def __init__(self, data_loader: ExperimentDataLoader, analyzer: ExperimentAnalyzer):
        """
        Initialize visualizer.
        Args:
            data_loader: ExperimentDataLoader
            analyzer: ExperimentAnalyzer
        """
        self.data_loader = data_loader
        self.analyzer = analyzer

    def _get_safe_filename(self, metric_name: str, suffix: str) -> str:
        """Generate safe file name (no spaces)."""
        safe_metric = metric_name.replace(' ', '').replace('　', '')
        return f'{safe_metric}_{suffix}'

    def plot_boxplots(self, save_path: Optional[str] = None):
        """
        Plot boxplots per metric.
        Args:
            save_path: save path; if None, show
        """
        metrics = self.data_loader.get_metric_names()
        n_metrics = len(metrics)
        
        # Dynamic subplot layout
        if n_metrics <= 2:
            fig, axes = plt.subplots(1, n_metrics, figsize=(9*n_metrics, 8))
        elif n_metrics <= 4:
            fig, axes = plt.subplots(2, 2, figsize=(18, 14))
            axes = axes.flatten()
        else:
            rows = (n_metrics + 2) // 3
            fig, axes = plt.subplots(rows, 3, figsize=(24, 6*rows))
            axes = axes.flatten()
        
        # Ensure iterable axes
        if n_metrics == 1:
            axes = [axes]
        
        # Ranking
        ranking = self.data_loader.get_algorithm_ranking()
        
        for idx, metric in enumerate(metrics):
            if idx >= len(axes):
                break
            ax = axes[idx]
            
            df = self.data_loader.get_data_for_metric(metric)
            
            if df.empty:
                ax.text(0.5, 0.5, 'No data', ha='center', va='center', fontsize=14)
                ax.set_title(f'{metric}')
                continue
            
            # Order by direction
            direction = self.data_loader.get_metric_direction(metric)
            algo_means = df.groupby('Algorithm')['Value'].mean()
            
            if direction == 'higher':
                algo_order = algo_means.sort_values(ascending=False).index.tolist()
            else:  # lower
                algo_order = algo_means.sort_values(ascending=True).index.tolist()
            
            # Limit to top 12 algorithms
            if len(algo_order) > 12:
                algo_order = algo_order[:12]
                df = df[df['Algorithm'].isin(algo_order)]
            
            # Palette
            colors = sns.color_palette("viridis", n_colors=len(algo_order))
            palette = {algo: colors[i] for i, algo in enumerate(algo_order)}
            
            # Boxplot
            sns.boxplot(data=df, x='Algorithm', y='Value', ax=ax, 
                       order=algo_order, palette=palette)
            
            ax.set_title(f'{metric} comparison', fontsize=14, fontweight='bold')
            ax.set_xlabel('')
            ax.set_ylabel('Value', fontsize=12)
            ax.tick_params(axis='x', rotation=45, labelsize=10)
            ax.grid(True, alpha=0.3, axis='y')
            
            # Statistical annotation
            friedman_result = self.analyzer.friedman_test(metric)
            if 'error' not in friedman_result:
                p_text = f"Friedman p = {friedman_result['p_value']:.4f}"
                if friedman_result['significant']:
                    p_text += " *"
                ax.text(0.02, 0.98, p_text, transform=ax.transAxes, 
                       fontsize=10, verticalalignment='top',
                       bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
        
        # Hide unused axes
        for idx in range(len(metrics), len(axes)):
            axes[idx].set_visible(False)
        
        # Title
        title_metrics = ", ".join(metrics)
        plt.suptitle(f'Overall performance across {title_metrics}', fontsize=16, fontweight='bold')
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            print(f"Boxplots saved to: {save_path}")
        else:
            plt.show()
        plt.close()
    
    def plot_scenario_performance(self, top_k: int = 6, save_path: Optional[str] = None):
        """
        Plot scenario performance (line charts).
        Args:
            top_k: top-k algorithms
            save_path: save path
        """
        metrics = self.data_loader.get_metric_names()
        n_metrics = len(metrics)
        
        # Dynamic subplot layout
        if n_metrics <= 2:
            fig, axes = plt.subplots(1, n_metrics, figsize=(9*n_metrics, 8))
        elif n_metrics <= 4:
            fig, axes = plt.subplots(2, 2, figsize=(18, 14))
            axes = axes.flatten()
        else:
            rows = (n_metrics + 2) // 3
            fig, axes = plt.subplots(rows, 3, figsize=(24, 6*rows))
            axes = axes.flatten()
        
        # Ensure iterable axes
        if n_metrics == 1:
            axes = [axes]
        
        # Ranking
        ranking = self.data_loader.get_algorithm_ranking()
        
        if ranking.empty:
            top_algorithms = self.data_loader.benchmark_algorithms[:top_k] if self.data_loader.benchmark_algorithms else []
        else:
            top_algorithms = ranking.head(top_k)['Algorithm'].tolist()
        
        if not top_algorithms:
            print("Warning: no algorithm data")
            return
        
        for idx, metric in enumerate(metrics):
            if idx >= len(axes):
                break
            ax = axes[idx]
            
            df = self.data_loader.get_data_for_metric(metric)
            
            if df.empty:
                ax.text(0.5, 0.5, 'No data', ha='center', va='center', fontsize=14)
                ax.set_title(f'{metric}')
                continue
            
            # Filter top-k
            df_top = df[df['Algorithm'].isin(top_algorithms)]
            
            # Mean per scenario
            pivot_df = df_top.groupby(['Scenario', 'Algorithm'])['Value'].mean().reset_index()
            pivot_df = pivot_df.pivot(index='Scenario', columns='Algorithm', values='Value')
            
            # Line plot
            colors = sns.color_palette("Set2", n_colors=len(top_algorithms))
            pivot_df.plot(marker='o', linewidth=2, ax=ax, color=colors)
            
            ax.set_title(f'{metric} - Top {top_k} across scenarios', fontsize=14, fontweight='bold')
            ax.set_xlabel('Scenario', fontsize=12)
            ax.set_ylabel('Value', fontsize=12)
            ax.legend(loc='best', fontsize=9)
            ax.grid(True, alpha=0.3)
            ax.tick_params(axis='x', rotation=45)
        
        # Hide unused axes
        for idx in range(len(metrics), len(axes)):
            axes[idx].set_visible(False)
        
        # Title
        title_metrics = ", ".join(metrics)
        plt.suptitle(f'Top-{top_k} stability across {title_metrics}', fontsize=16, fontweight='bold')
        plt.tight_layout()
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            print(f"Scenario performance saved to: {save_path}")
        else:
            plt.show()
        plt.close()
    
    def plot_radar_chart(self, top_k: int = 8, save_path: Optional[str] = None):
        """
        Plot radar chart (all metrics).
        Args:
            top_k: top-k algorithms
            save_path: save path
        """
        from math import pi
        
        ranking = self.data_loader.get_algorithm_ranking()
        
        if ranking.empty:
            print("No ranking data")
            return
        
        top_algorithms = ranking.head(top_k)['Algorithm'].tolist()
        metrics = self.data_loader.get_metric_names()
        
        # Normalized scores
        normalized_scores = []
        
        for metric in metrics:
            df = self.data_loader.get_data_for_metric(metric)
            direction = self.data_loader.get_metric_direction(metric)
            
            if not df.empty:
                algo_means = df.groupby('Algorithm')['Value'].mean()
                
                min_val = algo_means.min()
                max_val = algo_means.max()
                
                for algo in top_algorithms:
                    if algo in algo_means.index:
                        if max_val == min_val:
                            norm_score = 0.5
                        elif direction == 'higher':
                            norm_score = (algo_means[algo] - min_val) / (max_val - min_val)
                        else:  # lower
                            norm_score = 1 - (algo_means[algo] - min_val) / (max_val - min_val)
                    else:
                        norm_score = 0
                    
                    normalized_scores.append({
                        'Algorithm': algo,
                        'Metric': metric,
                        'Score': norm_score
                    })
        
        if not normalized_scores:
            print("No normalized data")
            return
        
        score_df = pd.DataFrame(normalized_scores)
        
        # Radar chart
        fig = plt.figure(figsize=(14, 12))
        ax = fig.add_subplot(111, projection='polar')
        
        angles = [n / len(metrics) * 2 * pi for n in range(len(metrics))]
        angles += angles[:1]
        
        # Palette
        colors = sns.color_palette("husl", n_colors=len(top_algorithms))
        
        for i, algo in enumerate(top_algorithms):
            algo_scores = score_df[score_df['Algorithm'] == algo].set_index('Metric')['Score']
            
            # Ensure metric order
            values = []
            for metric in metrics:
                if metric in algo_scores.index:
                    values.append(algo_scores[metric])
                else:
                    values.append(0)
            values += values[:1]
            
            ax.plot(angles, values, 'o-', linewidth=2, label=algo, color=colors[i])
            ax.fill(angles, values, alpha=0.1, color=colors[i])
        
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(metrics, fontsize=12)
        ax.set_yticks([0.2, 0.4, 0.6, 0.8, 1.0])
        ax.set_yticklabels(['0.2', '0.4', '0.6', '0.8', '1.0'], fontsize=10)
        ax.set_ylim(0, 1)
        ax.set_theta_offset(pi / 2)
        ax.set_theta_direction(-1)
        
        plt.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1), fontsize=11)
        
        # Title
        title_metrics = ", ".join(metrics)
        plt.title(f'Top-{top_k} radar chart ({title_metrics})', fontsize=16, fontweight='bold', pad=20)
        
        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')
            print(f"Radar chart saved to: {save_path}")
        else:
            plt.show()
        plt.close()
    
    def plot_performance_profiles(self, metric_name: str, save_path: Optional[str] = None):
        """
        Plot performance profile.
        Args:
            metric_name: metric name
            save_path: save path
        """
        profile_data = self.analyzer.calculate_performance_profile(metric_name)
        
        if profile_data.empty:
            print(f"Cannot plot performance profile for {metric_name}")
            return
        
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Ranking
        ranking = self.data_loader.get_algorithm_ranking()
        
        if not ranking.empty:
            # Top 8 by rank
            algo_order = ranking.head(8)['Algorithm'].tolist()
            profile_data = profile_data[profile_data['Algorithm'].isin(algo_order)]
        
        # Step plot
        algorithms = profile_data['Algorithm'].unique()
        colors = sns.color_palette("viridis", n_colors=len(algorithms))
        
        for i, algo in enumerate(algorithms):
            group = profile_data[profile_data['Algorithm'] == algo].sort_values('tau')
            ax.step(group['tau'], group['Probability'], where='post', 
                    linewidth=2, label=algo, color=colors[i])
        
        ax.set_xlabel('tau (performance ratio)', fontsize=12)
        ax.set_ylabel('P(ratio <= tau)', fontsize=12)
        ax.set_title(f'{metric_name} performance profile', fontsize=14, fontweight='bold')
        ax.legend(loc='best', fontsize=10)
        ax.grid(True, alpha=0.3)
        ax.set_xlim(profile_data['tau'].min(), profile_data['tau'].max())
        ax.set_ylim(0, 1.05)
        
        if save_path:
            from pathlib import Path
            save_path = Path(save_path)
            parent_dir = save_path.parent
            safe_filename = self._get_safe_filename(metric_name, 'performance_profile.png')
            safe_path = parent_dir / safe_filename
            
            plt.savefig(safe_path, dpi=300, bbox_inches='tight')
            print(f"Performance profile saved to: {safe_path}")
            
            # Remove empty file if path differs
            if save_path != safe_path and save_path.exists():
                try:
                    save_path.unlink()
                except:
                    pass
        else:
            plt.show()
        plt.close()

    def plot_heatmap(self, metric_name: str, save_path: Optional[str] = None):
        """Plot heatmap."""
        df = self.data_loader.get_data_for_metric(metric_name)
        
        if df.empty:
            print(f"No data for {metric_name}")
            return
        
        # Mean per scenario/algorithm
        pivot_df = df.groupby(['Scenario', 'Algorithm'])['Value'].mean().reset_index()
        pivot_df = pivot_df.pivot(index='Scenario', columns='Algorithm', values='Value')
        
        # Limit algorithms
        if pivot_df.shape[1] > 15:
            variances = pivot_df.var().sort_values(ascending=False)
            top_algorithms = variances.head(15).index.tolist()
            pivot_df = pivot_df[top_algorithms]
        
        # Colormap by direction
        direction = self.data_loader.get_metric_direction(metric_name)
        if direction == 'higher':
            cmap = 'YlOrRd'
        else:
            cmap = 'YlGnBu_r'
        
        fig, ax = plt.subplots(figsize=(16, 10))
        
        # Heatmap
        sns.heatmap(pivot_df, annot=True, fmt='.4f', cmap=cmap, 
            linewidths=0.5, ax=ax, cbar_kws={'label': 'Value'})
        
        ax.set_title(f'{metric_name} - performance heatmap', fontsize=14, fontweight='bold')
        ax.set_xlabel('Algorithm', fontsize=12)
        ax.set_ylabel('Scenario', fontsize=12)
        ax.tick_params(axis='x', rotation=45, labelsize=10)
        ax.tick_params(axis='y', labelsize=10)
        
        if save_path:
            from pathlib import Path
            save_path = Path(save_path)
            parent_dir = save_path.parent
            safe_filename = self._get_safe_filename(metric_name, 'heatmap.png')
            safe_path = parent_dir / safe_filename
            
            plt.savefig(safe_path, dpi=300, bbox_inches='tight')
            print(f"Heatmap saved to: {safe_path}")
            
            if save_path != safe_path and save_path.exists():
                try:
                    save_path.unlink()
                except:
                    pass
        else:
            plt.show()
        plt.close()
    
    def plot_statistical_significance(self, metric_name: str, save_path: Optional[str] = None):
        """Plot statistical significance matrix."""
        nemenyi_result = self.analyzer.nemenyi_posthoc(metric_name)
        
        if nemenyi_result.empty:
            print(f"Cannot plot significance matrix for {metric_name}")
            return
        
        # Limit algorithms
        if nemenyi_result.shape[0] > 12:
            ranking = self.data_loader.get_algorithm_ranking()
            if not ranking.empty:
                top_algorithms = ranking.head(12)['Algorithm'].tolist()
                # Ensure algorithms exist
                available_algorithms = [algo for algo in top_algorithms if algo in nemenyi_result.index]
                if available_algorithms:
                    nemenyi_result = nemenyi_result.loc[available_algorithms, available_algorithms]
        
        fig, ax = plt.subplots(figsize=(12, 10))
        
        # Significance mask
        mask = np.triu(np.ones_like(nemenyi_result, dtype=bool))
        
        # Heatmap
        sns.heatmap(nemenyi_result, mask=mask, annot=True, fmt='.4f', 
                cmap='RdYlGn_r', center=0.05, vmin=0, vmax=0.1,
                linewidths=0.5, ax=ax, cbar_kws={'label': 'p-value'})
        
        ax.set_title(f'{metric_name} - Nemenyi significance matrix', fontsize=14, fontweight='bold')
        ax.set_xlabel('Algorithm', fontsize=12)
        ax.set_ylabel('Algorithm', fontsize=12)
        ax.tick_params(axis='x', rotation=45, labelsize=10)
        ax.tick_params(axis='y', labelsize=10)
        
        # Significance note
        ax.text(0.02, 0.98, '* p < 0.05', transform=ax.transAxes, 
            fontsize=11, verticalalignment='top',
            bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))
        
        if save_path:
            from pathlib import Path
            save_path = Path(save_path)
            parent_dir = save_path.parent
            safe_filename = self._get_safe_filename(metric_name, 'significance.png')
            safe_path = parent_dir / safe_filename
            
            plt.savefig(safe_path, dpi=300, bbox_inches='tight')
            print(f"Significance matrix saved to: {safe_path}")
            
            if save_path != safe_path and save_path.exists():
                try:
                    save_path.unlink()
                except:
                    pass
        else:
            plt.show()
        plt.close()


class ExperimentReportGenerator:
    """Experiment report generator (HV/Coverage/Runtime)."""
    
    def __init__(self, data_loader: ExperimentDataLoader, 
                 analyzer: ExperimentAnalyzer,
                 visualizer: ExperimentVisualizer):
        """
        Initialize report generator.
        Args:
            data_loader: ExperimentDataLoader
            analyzer: ExperimentAnalyzer
            visualizer: ExperimentVisualizer
        """
        self.data_loader = data_loader
        self.analyzer = analyzer
        self.visualizer = visualizer
        
    def generate_full_report(self, output_dir: str = 'analysis_report'):
        """
        Generate full analysis report.
        Args:
            output_dir: output directory
        """
        # Project root
        if len(sys.path) > 0:
            project_root = os.path.dirname(sys.path[0])
        else:
            project_root = os.getcwd()
        
        # Output directory
        output_path = Path(os.path.join(project_root, output_dir))
        output_path.mkdir(parents=True, exist_ok=True)
        
        metrics = self.data_loader.get_metric_names()
        has_runtime = self.data_loader.has_runtime_metric()
        
        print("=" * 60)
        print("Generating full analysis report")
        print("=" * 60)
        print(f"Metrics detected: {', '.join(metrics)}")
        if has_runtime:
            print("Includes runtime (lower is better)")
        print("-" * 60)
        
        # 1. Summary
        print("\n[1/7] Computing statistical summary...")
        statistical_summary = self.analyzer.calculate_statistical_summary()
        
        # 2. Save summary to Excel
        print("[2/7] Saving statistical summary to Excel...")
        self._save_statistical_summary(statistical_summary, output_path / 'statistical_summary.xlsx')
        
        # 3. Boxplots
        print("[3/7] Generating boxplots...")
        self.visualizer.plot_boxplots(save_path=output_path / '01_boxplots.png')
        
        # 4. Scenario performance
        print("[4/7] Generating scenario performance charts...")
        self.visualizer.plot_scenario_performance(top_k=6, save_path=output_path / '02_scenario_performance.png')
        
        # 5. Radar chart
        print("[5/7] Generating radar chart...")
        self.visualizer.plot_radar_chart(top_k=8, save_path=output_path / '03_radar_chart.png')
        
        # 6. Per-metric analysis
        print("[6/7] Generating per-metric analysis...")
        
        for metric_name in metrics:
            metric_dir = output_path / metric_name
            metric_dir.mkdir(exist_ok=True)
            
            # Performance profile
            self.visualizer.plot_performance_profiles(
                metric_name, 
                save_path=metric_dir / f'{metric_name}_performance_profile.png'
            )
            
            # Heatmap
            self.visualizer.plot_heatmap(
                metric_name,
                save_path=metric_dir / f'{metric_name}_heatmap.png'
            )
            
            # Significance matrix
            self.visualizer.plot_statistical_significance(
                metric_name,
                save_path=metric_dir / f'{metric_name}_significance.png'
            )
        
        # 7. Text report
        print("[7/7] Generating text report...")
        self._generate_text_report(statistical_summary, output_path / 'analysis_report.txt')
        
        # 8. Algorithm stats
        print("[8/8] Generating algorithm statistics...")
        self._save_algorithm_stats(output_path / 'algorithm_statistics.csv')
        
        print("\n" + "=" * 60)
        print(f"Report generated. Files saved to: {output_path}")
        print("=" * 60)
        
    def _save_statistical_summary(self, summary: Dict, file_path: Path):
        """
        Save statistical summary to Excel.
        Args:
            summary: summary dict
            file_path: save path
        """
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Ranking
            if not summary['ranking'].empty:
                summary['ranking'].to_excel(writer, sheet_name='OverallRanking', index=False)
            
            # Best per metric
            best_df = pd.DataFrame(summary['best_algorithms']).T
            if not best_df.empty:
                best_df.to_excel(writer, sheet_name='BestByMetric')
            
            # Friedman test
            friedman_data = []
            for metric, result in summary['friedman_tests'].items():
                if 'error' not in result:
                    friedman_data.append({
                        'Metric': metric,
                        'FriedmanStat': result['statistic'],
                        'pValue': result['p_value'],
                        'Significant': 'Yes' if result['significant'] else 'No'
                    })
            
            if friedman_data:
                pd.DataFrame(friedman_data).to_excel(
                    writer, sheet_name='FriedmanTest', index=False
                )
            
            # Pairwise results
            pairwise_data = []
            for comp_name, comp_results in summary['pairwise_comparisons'].items():
                for metric, result in comp_results.items():
                    if 'error' not in result:
                        pairwise_data.append({
                            'Comparison': comp_name,
                            'Metric': metric,
                            'WilcoxonStat': result['statistic'],
                            'pValue': result['p_value'],
                            'Significant': 'Yes' if result['significant'] else 'No',
                            'EffectSize': result['effect_size'],
                            f'{result["algorithm1"]}Mean': result['mean1'],
                            f'{result["algorithm2"]}Mean': result['mean2'],
                            'Diff': result['diff']
                        })
            
            if pairwise_data:
                pd.DataFrame(pairwise_data).to_excel(
                    writer, sheet_name='Pairwise', index=False
                )
    
    def _save_algorithm_stats(self, file_path: Path):
        """Save algorithm statistics."""
        stats_df = self.data_loader.get_algorithm_stats()
        if not stats_df.empty:
            stats_df.to_csv(file_path, index=False, encoding='utf-8-sig')
            print(f"  Algorithm statistics saved to: {file_path}")
    
    def _generate_text_report(self, summary: Dict, file_path: Path):
        """
        Generate text report.
        Args:
            summary: summary dict
            file_path: save path
        """
        metrics = self.data_loader.get_metric_names()
        has_runtime = self.data_loader.has_runtime_metric()
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("Multi-objective DJSP - RL experiment report\n")
            f.write("=" * 80 + "\n\n")
            
            # 1. Overview
            f.write("1. Overview\n")
            f.write("-" * 40 + "\n")
            f.write(f"  Scenarios: {len(self.data_loader.raw_data)}\n")
            f.write(f"  Algorithms: {len(self.data_loader.benchmark_algorithms) if self.data_loader.benchmark_algorithms else 0}\n")
            f.write(f"  Metrics: {len(metrics)} ({', '.join(metrics)})\n")
            if has_runtime:
                f.write("  Runtime: lower is better\n")
            f.write(f"  Records: {len(self.data_loader.processed_data) if self.data_loader.processed_data is not None else 0}\n\n")
            
            # 2. Ranking
            f.write("2. Overall Ranking\n")
            f.write("-" * 40 + "\n")
            if not summary['ranking'].empty:
                for _, row in summary['ranking'].iterrows():
                    f.write(f"  {int(row['Rank'])}. {row['Algorithm']}: {row['Normalized']:.4f}\n")
            else:
                f.write("  No ranking data\n")
            f.write("\n")
            
            # 3. Best by metric
            f.write("3. Best by Metric\n")
            f.write("-" * 40 + "\n")
            if summary['best_algorithms']:
                for metric, best_info in summary['best_algorithms'].items():
                    direction = "higher is better" if best_info['direction'] == 'higher' else "lower is better"
                    f.write(f"  {metric}: {best_info['algorithm']} ({best_info['value']:.4f}, {direction})\n")
            else:
                f.write("  No best-algorithm data\n")
            f.write("\n")
            
            # 4. Statistical tests
            f.write("4. Statistical Tests\n")
            f.write("-" * 40 + "\n")
            
            for metric, result in summary['friedman_tests'].items():
                if 'error' not in result:
                    f.write(f"\n  {metric}:\n")
                    f.write(f"    Friedman: p = {result['p_value']:.4e} ")
                    f.write("(significant)\n" if result['significant'] else "(not significant)\n")
                    
                    if result['significant'] and 'nemenyi' in result:
                        nemenyi = result['nemenyi']
                        if not nemenyi.empty:
                            f.write("    Nemenyi pairs (p < 0.05):\n")
                            sig_pairs = 0
                            for i, algo1 in enumerate(nemenyi.columns):
                                for j, algo2 in enumerate(nemenyi.columns):
                                    if i < j and nemenyi.iloc[i, j] < 0.05:
                                        f.write(f"      - {algo1} vs {algo2}: p = {nemenyi.iloc[i, j]:.4f}\n")
                                        sig_pairs += 1
                            if sig_pairs == 0:
                                f.write("      No significant pairs\n")
            
            # 5. Conclusions
            f.write("\n5. Conclusions\n")
            f.write("-" * 40 + "\n")
            
            if not summary['ranking'].empty:
                best_algo = summary['ranking'].iloc[0]['Algorithm']
                second_algo = summary['ranking'].iloc[1]['Algorithm'] if len(summary['ranking']) > 1 else None
                
                f.write(f"\n  Best overall: {best_algo}\n")
                f.write(f"\n  Advantages:\n")
                
                # Metrics where best_algo is best
                best_algo_advantages = []
                for metric, best_info in summary['best_algorithms'].items():
                    if best_info['algorithm'] == best_algo:
                        best_algo_advantages.append(metric)
                
                if best_algo_advantages:
                    for metric in best_algo_advantages:
                        f.write(f"    - Best on {metric}\n")
                else:
                    f.write(f"    - Highest overall score, not best on any single metric\n")
                
                if second_algo:
                    f.write(f"\n  Runner-up: {second_algo}\n")
                
                f.write("\n  Recommendations:\n")
                f.write(f"    - For overall performance: {best_algo}\n")
                
                # Metric-specific picks
                hv_best = summary['best_algorithms'].get('HV', {}).get('algorithm', '')
                coverage_best = summary['best_algorithms'].get('Coverage', {}).get('algorithm', '')
                runtime_best = summary['best_algorithms'].get('Runtime', {}).get('algorithm', '')
                
                if hv_best:
                    f.write(f"    - For convergence (HV): {hv_best}\n")
                if coverage_best:
                    f.write(f"    - For diversity (Coverage): {coverage_best}\n")
                if runtime_best:
                    f.write(f"    - For efficiency (Runtime): {runtime_best}\n")
            else:
                f.write("\n  No conclusions: no ranking data\n")
            
            f.write("\n" + "=" * 80 + "\n")
            f.write("Generated: " + pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S") + "\n")
            f.write("=" * 80 + "\n")

    def generate_full_report_with_detailed_analysis(self, output_dir: str = 'analysis_report'):
        """
        Generate full report with detailed analysis.
        """
        # Reuse base report generation
        self.generate_full_report(output_dir)
        
        output_path = Path(output_dir)
        
        # ==================== 1. Detailed scenario metrics ====================
        print("\n[Enhanced 1/3] Generating detailed scenario metrics...")
        detailed_excel_path = output_path / '01_detailed_scenario_metrics.xlsx'
        self.data_loader.generate_detailed_scenario_metrics(detailed_excel_path)
        
        # ==================== 2. Win-rate stats ====================
        print("[Enhanced 2/3] Generating win-rate statistics...")
        win_rate_df = self.data_loader.generate_win_rate_statistics()
        if not win_rate_df.empty:
            win_rate_path = output_path / '02_win_rate_statistics.csv'
            win_rate_df.to_csv(win_rate_path, index=False, encoding='utf-8-sig')
            print(f"  Win-rate stats saved to: {win_rate_path}")
            
            # Also save Excel
            win_rate_excel_path = output_path / '02_win_rate_statistics.xlsx'
            with pd.ExcelWriter(win_rate_excel_path, engine='openpyxl') as writer:
                win_rate_df.to_excel(writer, sheet_name='WinRate', index=False)
        
        # ==================== 3. Stability-weighted ranking ====================
        print("[Enhanced 3/3] Generating stability-weighted ranking...")
        stability_ranking = self.data_loader.get_algorithm_ranking_with_stability(stability_weight=0.3)
        if not stability_ranking.empty:
            ranking_path = output_path / '03_stability_weighted_ranking.csv'
            stability_ranking.to_csv(ranking_path, index=False, encoding='utf-8-sig')
            print(f"  Stability-weighted ranking saved to: {ranking_path}")
            
            # Compare multiple weights
            weights = [0.1, 0.2, 0.3, 0.4, 0.5]
            ranking_comparison = []
            
            for w in weights:
                rank_df = self.data_loader.get_algorithm_ranking_with_stability(stability_weight=w)
                if not rank_df.empty:
                    for _, row in rank_df.iterrows():
                        ranking_comparison.append({
                            'StabilityWeight': w,
                            'Rank': row['Rank'],
                            'Algorithm': row['Algorithm'],
                            'OverallScore': row['OverallScore']
                        })
            
            if ranking_comparison:
                comp_df = pd.DataFrame(ranking_comparison)
                comp_pivot = comp_df.pivot(index='Algorithm', columns='StabilityWeight', values='OverallScore')
                comp_pivot.to_csv(output_path / '03_stability_weight_sensitivity.csv', 
                                encoding='utf-8-sig')
        
        print("\n" + "=" * 60)
        print("Enhanced analysis complete. Detailed reports generated.")
        print("=" * 60)


# ==================== Usage example ====================

def run_complete_analysis(data_path: str = 'test_result', 
                         sub_dir: str = 'ablation',
                         output_dir: str = 'analysis_report'):
    """
    Run the full analysis pipeline.
    Args:
        data_path: data root
        sub_dir: subdirectory
        output_dir: output directory
    """
    print("=" * 60)
    print("Multi-objective DJSP - RL analysis system")
    print("=" * 60)
    
    # 1. Load data
    print("\n[Stage 1] Loading experiment data...")
    loader = ExperimentDataLoader(data_path)
    loader.load_all_experiments(sub_dir=sub_dir)
    
    if not loader.raw_data:
        print("Error: no data loaded; check data path")
        return None, None, None, None
    
    # 2. Analyzer
    print("\n[Stage 2] Initializing analyzer...")
    analyzer = ExperimentAnalyzer(loader)
    
    # 3. Visualizer
    print("\n[Stage 3] Initializing visualizer...")
    visualizer = ExperimentVisualizer(loader, analyzer)
    
    # 4. Report
    print("\n[Stage 4] Generating full report...")
    reporter = ExperimentReportGenerator(loader, analyzer, visualizer)
    reporter.generate_full_report(output_dir=output_dir)
    
    return loader, analyzer, visualizer, reporter


def data_analysis_report(excel_path = 'test_result', sub_path= 'Ref_Learning'):
    # ==================== Project root ====================
    project_root = sys.path[0]
    print(f"Project root: {project_root}")
    
    # ==================== Absolute paths ====================
    base_output_path = os.path.join(project_root, "test_result", "experimental_analysis", sub_path)
    print(f"Base output dir: {base_output_path}")
    
    # Ensure directory exists
    os.makedirs(base_output_path, exist_ok=True)
    
    # Absolute paths for outputs
    detailed_metrics_path = os.path.join(base_output_path, "detailed_metrics.xlsx")
    win_rate_path = os.path.join(base_output_path, "win_rate_statistics.xlsx")
    stability_ranking_path = os.path.join(base_output_path, "stability_weighted_ranking.xlsx")
    
    print(f"Detailed report path: {detailed_metrics_path}")
    
    # 1. Basic analysis
    loader, analyzer, visualizer, reporter = run_complete_analysis(
        data_path=excel_path,
        sub_dir=sub_path,
        output_dir=base_output_path  # Absolute path
    )
    
    # 2. Enhanced analysis
    if loader and loader.processed_data is not None:
        print("\n" + "=" * 60)
        print("Starting enhanced analysis...")
        print("=" * 60)
        
        # 2.1 Detailed scenario metrics
        loader.generate_detailed_scenario_metrics(detailed_metrics_path)
        print(f"Detailed report saved to: {os.path.abspath(detailed_metrics_path)}")
        
        # 2.2 Win-rate stats
        win_rate = loader.generate_win_rate_statistics()
        if not win_rate.empty:
            win_rate.to_excel(win_rate_path, index=False)
            print(f"Win-rate stats saved to: {os.path.abspath(win_rate_path)}")
        
        # 2.3 Stability-weighted ranking
        stability_ranking = loader.get_algorithm_ranking_with_stability(stability_weight=0.3)
        if not stability_ranking.empty:
            stability_ranking.to_excel(stability_ranking_path, index=False)
            print(f"Stability-weighted ranking saved to: {os.path.abspath(stability_ranking_path)}")


if __name__ == "__main__":
    # Usage example
    data_analysis_report(excel_path='test_result', sub_path='Ref_Learning')