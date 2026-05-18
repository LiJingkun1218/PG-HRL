# shared_modules.py
import numpy as np
import sys
sys.path
import os
import itertools
from openpyxl import load_workbook,Workbook
import warnings
import random 
import copy
from scipy.optimize import minimize
import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
from agent import job_creation, agent_machine, sequencing 
from common.cfunctions import (state_multi_channel, sequencing_data_generation, build_experience)
import common.multiobject as mutilobjectivemanager
warnings.filterwarnings("ignore")


class ShopFloor:
    """Unified shop-floor simulation."""
    def __init__(self, env, job_numbers, parameters, brain_machine, train=False,**kwargs):
        self.train = train
        self.env = env        
        self.m_no = parameters['machine_count']
        self.use_ratio = parameters['utilization'] 
        self.tightness = parameters['tightness_factor']
        self.h_t = parameters['process_time_range'][1]
        self.l_t = parameters['process_time_range'][0]
        self.job_numbers = job_numbers
        self.m_list = [] 
        self.job_objectives_records = []  # Job-level multi-objective records
        self.preference_vector = kwargs.get('preference_vector', None)  # Preference vector
        self.ablation = kwargs.get('ablation', None)
        
        # Create machines
        for i in range(self.m_no):
            expr1 = f'''self.m_{i} = agent_machine.machine(env, {i}, print=0)'''
            exec(expr1)
            expr2 = f'''self.m_list.append(self.m_{i})'''
            exec(expr2)
            
        # Create job generator
        if 'seed' in kwargs:
            self.job_creator = job_creation.creation\
            (self.env, self.job_numbers, self.m_list, [self.l_t, self.h_t], self.tightness, self.use_ratio, train, seed=kwargs['seed'])
        else:
           self.job_creator = job_creation.creation(
            self.env, self.job_numbers, self.m_list, [self.l_t, self.h_t], self.tightness, self.use_ratio, train, random_seed=True
            )

        # Initialize machines
        for i, m in enumerate(self.m_list):
            m.initialization(self.m_list, self.job_creator)
            
        # Scheduling rule setup
        if 'sequencing_rule' in kwargs: 
            self._setup_scheduling_rule(kwargs['sequencing_rule'])
        else:      
            self._setup_neural_network(brain_machine,kwargs)
    
    def _setup_neural_network(self, brain_machine,kwargs=None):        
        address = kwargs.get('address', None)
        if self.train == False:
            self.sqc_brain = brain_machine.sequencing_brain(self.env, self.job_creator, 
                self.m_list, self.job_numbers, ma_no=self.m_no,tightness=self.tightness, 
                address=address,ablation=self.ablation, preference_vector=self.preference_vector
            )            
        else:
            self.sqc_brain = brain_machine.sequencing_brain(self.env, self.job_creator, 
                self.m_list, self.job_numbers, ma_no=self.m_no, tightness=self.tightness,
                address=address,ablation=self.ablation, preference_vector=self.preference_vector,
                train=True
            )            
            for m in self.m_list:
                m.sqc_brain = self.sqc_brain
        self.job_creator.sqc_brain = self.sqc_brain

    def _setup_scheduling_rule(self, rule_name):        
        for m in self.m_list:
            order = f"m.job_sequencing = sequencing.{rule_name}"
            try:
                exec(order)
            except:
                print(f"Rule assigned to machine {m.m_idx} is invalid!")
                raise Exception  
    
    def simulation(self,bit = 0):
        """Run simulation.""" 
        # bit = 0: per-operation optimization; bit = 1: global optimization
        self.env.run()
        if bit == 0:
            self.collect_job_level_objectives()
        else:
            self.collect_total_objectives()
       
    def collect_job_level_objectives(self):
        """Collect per-job objectives."""
        self.job_objectives_records = []
        for job_id, record in self.job_creator.production_record.items():            
            objective1 = record[1]; objective2 = record[2];  objective3 = record[3]             
            job_record = {
                'job_id': job_id,
                'arrival_order': job_id,  # Use job_id as arrival order
                'objectives': np.array([objective1, objective2, objective3])
            }            
            self.job_objectives_records.append(job_record)
        self.job_objectives_records.sort(key=lambda x: x['arrival_order'])

    def collect_total_objectives(self):
        """Collect total objectives (summed)."""
        self.job_objectives_records = []
        objective1 = 0 ;objective2 = 0; objective3 = 0
        for job_id, record in self.job_creator.production_record.items():            
            objective1 += record[1]; objective2 += record[2];  objective3 += record[3]             
        job_record = {                
            'objectives': np.array([objective1, objective2, objective3])
        }            
        self.job_objectives_records.append(job_record)


class MultiObjectiveManager:
    """Multi-objective analysis manager (HV and coverage only)."""
    def __init__(self):
        self.all_experiment_data = {}           # All experiment data
        self.global_ideal = None                 # Global ideal point
        self.global_nadir = None                  # Global nadir point
        self.normalization_ranges = None          # Normalization ranges
        # Algorithm Pareto sets
        self.algorithm_pareto_solutions = {}      # Non-dominated per algorithm
        self.global_pareto_front = None           # Global Pareto front
    
    def add_experiment_data(self, rule_name, run_id, job_objectives):
        """Add experiment data."""
        if rule_name not in self.all_experiment_data:
            self.all_experiment_data[rule_name] = {}
        self.all_experiment_data[rule_name][run_id] = job_objectives
        # Clear caches
        self.algorithm_pareto_solutions = {}
        self.global_pareto_front = None
    
    def calculate_global_reference_points(self):
        """Compute global reference points for normalization."""
        all_solutions = []
        for rule_data in self.all_experiment_data.values():
            for run_data in rule_data.values():
                for job_data in run_data:
                    if 'objectives' in job_data:
                        all_solutions.append(job_data['objectives'])
        
        if all_solutions:
            all_solutions_array = np.array(all_solutions)
            self.global_ideal = np.min(all_solutions_array, axis=0)
            self.global_nadir = np.max(all_solutions_array, axis=0)
            self.normalization_ranges = self.global_nadir - self.global_ideal
            self.normalization_ranges[self.normalization_ranges == 0] = 1.0
        else:
            self.global_ideal = np.array([0, 0, 0])
            self.global_nadir = np.array([1, 1, 1])
            self.normalization_ranges = np.array([1, 1, 1])
    
    def normalize_solutions(self, solutions):
        """Normalize solutions to [0,1]."""
        if self.global_ideal is None:
            self.calculate_global_reference_points()
        
        solutions_array = np.array(solutions)
        normalized = (solutions_array - self.global_ideal) / self.normalization_ranges
        return np.clip(normalized, 0, 1)
    
    def _calculate_pareto_mask(self, solutions):
        """Compute Pareto dominance mask (minimization)."""
        n = len(solutions)
        if n == 0:
            return np.array([], dtype=bool)
        
        mask = np.ones(n, dtype=bool)
        for i in range(n):
            if not mask[i]:
                continue
            for j in range(n):
                if i != j and mask[j]:
                    if np.all(solutions[j] <= solutions[i]) and np.any(solutions[j] < solutions[i]):
                        mask[i] = False
                        break
        return mask
    
    def calculate_pareto_fronts(self):
        """Compute per-algorithm non-dominated sets and global front."""
        self.calculate_global_reference_points()
        
        # Initialize
        self.algorithm_pareto_solutions = {}
        all_solutions_with_info = []
        
        # Non-dominated per algorithm
        for rule_name, rule_data in self.all_experiment_data.items():
            algorithm_solutions = []
            algorithm_info = []
            
            for run_id, run_data in rule_data.items():
                for job_data in run_data:
                    if 'objectives' in job_data:
                        algorithm_solutions.append(job_data['objectives'])
                        algorithm_info.append({
                            'rule_name': rule_name,
                            'objectives': job_data['objectives']
                        })
            
            if algorithm_solutions:
                solutions_array = np.array(algorithm_solutions)
                pareto_mask = self._calculate_pareto_mask(solutions_array)
                
                self.algorithm_pareto_solutions[rule_name] = {
                    'solutions': solutions_array[pareto_mask],
                    'info': [info for i, info in enumerate(algorithm_info) if pareto_mask[i]]
                }
                
                # Collect for global front
                all_solutions_with_info.extend([
                    (sol, info) for sol, info in zip(solutions_array[pareto_mask], 
                                                      [info for i, info in enumerate(algorithm_info) if pareto_mask[i]])
                ])
            else:
                self.algorithm_pareto_solutions[rule_name] = {
                    'solutions': np.array([]),
                    'info': []
                }
        
        # Global Pareto front
        if all_solutions_with_info:
            all_solutions = np.array([item[0] for item in all_solutions_with_info])
            all_info = [item[1] for item in all_solutions_with_info]
            
            global_pareto_mask = self._calculate_pareto_mask(all_solutions)
            
            self.global_pareto_front = {
                'solutions': all_solutions[global_pareto_mask],
                'info': [info for i, info in enumerate(all_info) if global_pareto_mask[i]]
            }
        else:
            self.global_pareto_front = {'solutions': np.array([]), 'info': []}
    
    def calculate_hypervolume(self, solutions):
        """Compute hypervolume."""
        if len(solutions) < 1:
            return 0
        
        # Normalize
        normalized = self.normalize_solutions(solutions)
        fixed_ref_point = np.ones(3) * 1.01
        
        try:
            from pymoo.indicators.hv import HV
            ind = HV(ref_point=fixed_ref_point)
            return ind(normalized)
        except ImportError:
            # Simplified fallback
            if len(normalized) == 0:
                return 0
            
            solutions_sorted = normalized[np.argsort(normalized[:, 0])]
            hv_value = 0.0
            
            for sol in solutions_sorted:
                volume = 1.0
                for i in range(len(sol)):
                    if fixed_ref_point[i] > sol[i]:
                        volume *= (fixed_ref_point[i] - sol[i])
                    else:
                        volume = 0
                        break
                if volume > hv_value:
                    hv_value = volume
            
            return hv_value
    
    def calculate_coverage(self):
        """
        Compute coverage.
        Coverage = (algo solutions in global front / total global front) * 100%
        """
        if self.global_pareto_front is None:
            self.calculate_pareto_fronts()
        
        coverage = {}
        
        if self.global_pareto_front and len(self.global_pareto_front['info']) > 0:
            total_global = len(self.global_pareto_front['info'])
            
            # Count contributions per algorithm
            for rule_name in self.algorithm_pareto_solutions.keys():
                count = sum(1 for info in self.global_pareto_front['info'] 
                           if info['rule_name'] == rule_name)
                coverage[rule_name] = (count / total_global) * 100
        else:
            for rule_name in self.algorithm_pareto_solutions.keys():
                coverage[rule_name] = 0.0
        
        return coverage
    
    def get_metrics(self):
        """
        Return hypervolume and coverage metrics.
        Returns: (hypervolumes, coverages)
        """
        if not self.algorithm_pareto_solutions:
            self.calculate_pareto_fronts()
        
        # Hypervolume
        hypervolumes = {}
        for rule_name, data in self.algorithm_pareto_solutions.items():
            hypervolumes[rule_name] = self.calculate_hypervolume(data['solutions'])
        
        # Coverage
        coverages = self.calculate_coverage()
        
        return hypervolumes, coverages
    
    def save_to_excel(self, scenario_id, cyc, benchmark, hypervolumes, coverages, cpath='ablation'):
        """Save hypervolume and coverage to Excel."""
        
        excel_path = f"{sys.path[0]}\\test_result\\{cpath}\\{scenario_id}_metrics.xlsx"
        
        # Ensure directory exists
        directory = os.path.dirname(excel_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)
        
        try:
            wb = load_workbook(excel_path)
        except FileNotFoundError:
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
        
        # Metric config
        metrics = [
            {'name': 'Hypervolume (HV)', 'data': hypervolumes, 'format': '{:.6f}'},
            {'name': 'Coverage', 'data': coverages, 'format': '{:.2f}%'}
        ]
        
        for metric in metrics:
            sheet_name = metric['name']
            data_dict = metric['data']
            
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
                headers = ['Scenario', 'Run'] + benchmark
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
            
            new_row = [scenario_id, cyc + 1]
            
            for algo in benchmark:
                value = data_dict.get(algo, 0)
                new_row.append(value)
            
            next_row = ws.max_row + 1
            for col_idx, value in enumerate(new_row, 1):
                ws.cell(row=next_row, column=col_idx, value=value)
        
        wb.save(excel_path)
        wb.close()


# ========== PD-MORL core components ==========
class PD_MORL_PreferenceManager:
    """PD-MORL preference manager (preference space + HER)."""
    def __init__(self, reward_size=3, step_size=0.2):
        self.reward_size = reward_size
        self.w_batch = self.generate_preference_batch(step_size)
        self.her_samples = 3  # HER samples per transition
    
    def generate_preference_batch(self, step_size):
        """Generate preference space w_batch."""
        mesh_array = []
        for i in range(self.reward_size):
            mesh_array.append(np.arange(0, 1 + step_size, step_size))
        
        w_batch = np.array(list(itertools.product(*mesh_array)))
        w_batch = w_batch[w_batch.sum(axis=1) == 1, :]
        return np.unique(w_batch, axis=0)
    
    def sample_preference(self):
        """Sample from preference space."""
        idx = random.randint(0, len(self.w_batch) - 1)
        return torch.tensor(self.w_batch[idx], dtype=torch.float32)
    
    def get_her_preferences(self, original_preference, achieved_objectives):
        """Generate HER preferences for each transition."""
        her_preferences = []
        
        # 1) Original preference
        her_preferences.append(original_preference)
        
        # 2) Random preferences (HER)
        for _ in range(self.her_samples - 1):
            random_pref = torch.tensor(np.random.dirichlet([1, 1, 1]), dtype=torch.float32)
            her_preferences.append(random_pref)
        
        return her_preferences


class ContinuousSequencingBrain:
    """Base class for sequencing brain."""
    def __init__(self, env, job_creator, all_machines, job_numbers, *args, **kwargs):
        self.env = env  # SimPy environment
        self.job_creator = job_creator  # Job generator
        self.m_list = all_machines  # Machine list
        self.m_no = len(self.m_list)  # Machine count
        self.warm_up = 0.05 * job_numbers * job_creator.avg_pt  # Warm-up time
        self.span = job_numbers * job_creator.avg_pt  # Total sim time
        self.job_creator.build_sqc_experience_repository(self.m_list)  # Replay buffer
        self.trajectory_buffer = []  # Trajectory buffer
        self.rep_memo = []  # Replay memory
        self.discount_factor = 0.99  # Discount factor
        self.epsilon = 0.1  # Exploration rate
        self.loss_record = []  # Loss record
        self.rule_loss_record = []  # Policy loss record
        self.value_loss_record = []  # Value loss record
        self.sample_reward_record = []  # Reward record
        self.train_reward_record = []        
        self.current_preference_vector = kwargs.get('preference_vector', None)
        self.train = kwargs.get('train', False)    
        self.address = kwargs.get('address', None)
        self.ablation = kwargs.get('ablation', None)        
        self.gae_lambda = 0.95  # GAE lambda
        self.entropy_coeff = 0.01  # Entropy coeff
        self.value_loss_coeff = 0.5  # Value loss coeff
        self.max_grad_norm = 0.5  # Gradient clip
        self.trajectory_buffer_size = 1280  # Min trajectory length
        self.cosine_weight = kwargs.get('cosine_weight', 0.1)
        self.samples = 0

        # Training params (override in subclasses)
        self.training_epochs = None
        self.network = None
        self.algorithm_name = None
        
        self.pd_morl_manager =  PD_MORL_PreferenceManager(
            reward_size=3,
            step_size=kwargs.get('preference_step', 0.2)
        )

        for m in self.m_list:           
            m.job_sequencing = self.action_default  # Default sequencing
        
        self.func_list = [sequencing.CR, sequencing.MDD, sequencing.MOD, sequencing.MS, sequencing.ATC, sequencing.EDD]  # Candidate rules
        self.multi_obj_manager = mutilobjectivemanager.MultiObjectiveManager(num_objectives=3)    
        self.training_step_count = 0
        
        if self.train == True: 
            self.input_size = len(state_multi_channel(self, sequencing_data_generation(self.m_list[0]))) 
            self.env.process(self.warm_up_process())
            
        if self.train == False:         
            self.input_size = len(state_multi_channel(self, sequencing_data_generation(self.m_list[0])))
            self.multi_obj_manager.get_per_baselines(self.address)
            for m in self.m_list:
                m.job_sequencing = self.action_sqc_rule  
    
    def warm_up_process(self):  # Warm-up
        for m in self.m_list:
             m.job_sequencing = self.action_warm_up
        for idx, func in enumerate(self.func_list):            
            self.func_selection = idx
            print('Time {}: rule set to {}'.format(self.env.now, func))
            yield self.env.timeout(int(self.warm_up / (len(self.func_list) + 1)))  
            
        for m in self.m_list:
            m.job_sequencing = self.action_default 
            
        print("Time {} to {}: random rules for all machines.".format(self.env.now, self.warm_up))
        yield self.env.timeout(self.warm_up - self.env.now - 1)        
        
        self.rep_memo = copy.deepcopy(self.trajectory_buffer)

        print("Time {}: machines start RL-based sequencing.".format(self.env.now))
        for m in self.m_list:
            m.job_sequencing = self.action_sqc_rule 
            
    # Default action phase
    def action_default(self, sqc_data):
        m_idx = sqc_data[-1]
        current_queue = sqc_data[-2]
        queue_size = len(current_queue)
        if queue_size > 1:
            s_t = state_multi_channel(self, sqc_data)
            job_weights = self.multi_obj_manager.calculate_importance(self.job_creator, current_queue)
            preference = self.multi_obj_manager.calculate_preference(max(current_queue))
            job_position, _ = sequencing.FIFO(sqc_data)      
            a_rule = torch.tensor(preference, dtype=torch.float32)  
            _, reward = self.selection_job_reward( job_weights, preference)             
            build_experience(self, self.env.now, m_idx, s_t, a_rule, reward, preference,job_weights)       
        else:
            job_position = 0
        
        delay = self.calculate_delay(sqc_data, job_position) if not (self.ablation == "Ablation1") else 0            
            
        return job_position, delay 
      
    # Warm-up action phase
    def action_warm_up(self, sqc_data):       
        current_queue = sqc_data[-2]
        queue_size = len(current_queue)
        m_idx = sqc_data[-1]
        if queue_size > 1:
            s_t = state_multi_channel(self, sqc_data) 
            job_position, _ = self.func_list[self.func_selection](sqc_data)            
            job_weights = self.multi_obj_manager.calculate_importance(self.job_creator, current_queue)
            preference = self.multi_obj_manager.calculate_preference(max(current_queue))   
            a_rule = torch.tensor(preference, dtype=torch.float32) 
            _, reward = self.selection_job_reward(job_weights, preference)  
            build_experience(self, self.env.now, m_idx, s_t, a_rule, reward, preference,job_weights)
        else:
            job_position = 0
        
        delay = self.calculate_delay(sqc_data, job_position) if not (self.ablation == "Ablation1") else 0             
        
        return job_position, delay  
  
    # RL action phase
    def action_sqc_rule(self, sqc_data):
        current_queue = sqc_data[-2]   
        queue_size = len(current_queue)
        m_idx = sqc_data[-1]
        if queue_size > 1:
            s_t = state_multi_channel(self, sqc_data) 
            s_t_reshaped = s_t.reshape([1, self.input_size]) 
            if self.train == False:  # Inference
                preference = self.current_preference_vector                
            else:  # Training
                preference = self.multi_obj_manager.calculate_preference(max(current_queue))
                if self.ablation == "Ablation3":
                    preference = torch.tensor(np.array(np.random.dirichlet([1, 1, 1])), dtype=torch.float32)
                if self.ablation == "Ablation4":
                    preference = self.pd_morl_manager.sample_preference() 
            job_weights = self.multi_obj_manager.calculate_importance(self.job_creator, current_queue)
                
            if random.random() < self.epsilon or self.ablation == "Ablation3":                
                order_coeff = torch.tensor(np.random.dirichlet([1, 1, 1]), dtype=torch.float32)             
            else:   
                order_coeff = self.network.forward(s_t_reshaped, preference.reshape(1, 3))
                order_coeff = order_coeff.squeeze(0)
            
            job_position, reward = self.selection_job_reward( job_weights, order_coeff)
            
            if self.train == True:
                if self.ablation == "Ablation4":
                    # Cosine similarity between output and preference
                    cosine_sim = F.cosine_similarity(
                        order_coeff.unsqueeze(0),
                        preference.unsqueeze(0)
                    ).item()
                    # HER preference samples
                    her_preferences = self.pd_morl_manager.get_her_preferences(preference, job_weights[job_position, :])
                
                    # Store experience per preference
                    for her_pref in her_preferences:
                        # Recompute reward for preference
                        _, her_reward = self.selection_job_reward(job_weights, her_pref)
                        if torch.allclose(her_pref, preference):
                            final_reward = her_reward * (1.0 + self.cosine_weight * max(0, cosine_sim))
                        else:
                            final_reward = her_reward
                            
                    build_experience(self, self.env.now, m_idx, s_t, order_coeff.detach(), final_reward, preference, job_weights)                    
                else:
                    build_experience(self, self.env.now, m_idx, s_t, order_coeff.detach(), reward, preference, job_weights)
            self.samples += 1

        else:
            job_position = 0    

        delay = self.calculate_delay( sqc_data, job_position) if not (self.ablation == "Ablation1")  else 0             
        
        return job_position, delay
 
#    # ========== Integrated PD-MORL ==========
#     def action_sqc_rule(self, sqc_data):
#         """Integrated PD-MORL (three components)."""
#         current_queue = sqc_data[-2]
#         queue_size = len(current_queue)
#         m_idx = sqc_data[-1]
        
#         if queue_size > 1:
#             s_t = state_multi_channel(self, sqc_data)
#             s_t_reshaped = s_t.reshape([1, self.input_size])
            
#             # Job importance (original reward)
#             job_weights = self.multi_obj_manager.calculate_importance(self.job_creator, current_queue)
            
#             if self.train == False:  # Inference
#                 # Keep original logic
#                 preference = self.current_preference_vector 
#             else:  # Training
#                 # ========== PD-MORL 1: preference-driven update ==========
#                 preference = self.pd_morl_manager.sample_preference()    
            
#             # Action selection
#             if random.random() < self.epsilon :
#                 order_coeff = torch.tensor(np.random.dirichlet([1, 1, 1]), dtype=torch.float32)
#             else:
#                 # ========== PD-MORL: network takes (state, preference) ==========
#                 order_coeff = self.network.forward(s_t_reshaped, preference.reshape(1, 3))
#                 order_coeff = order_coeff.squeeze(0)
            
#             # ========== Original reward kept ==========
#             job_position, original_reward = self.selection_job_reward(job_weights, order_coeff)
            
#             # ========== PD-MORL 2: cosine similarity term ==========
#             if self.train:
#                 # Cosine similarity
#                 cosine_sim = F.cosine_similarity(
#                     order_coeff.unsqueeze(0),
#                     preference.unsqueeze(0)
#                 ).item()
#                 # HER preference samples
#                 her_preferences = self.pd_morl_manager.get_her_preferences(preference, job_weights[job_position, :])
                
#                 # Store experience per preference
#                 for her_pref in her_preferences:
#                     # Recompute reward for preference
#                     _, her_reward = self.selection_job_reward(job_weights, her_pref)
#                     if torch.allclose(her_pref, preference):
#                         final_reward = her_reward * (1.0 + self.cosine_weight * max(0, cosine_sim))
#                     else:
#                         final_reward = her_reward
                        
#                 build_experience(self, self.env.now, m_idx, s_t, order_coeff.detach(), final_reward, preference, job_weights)                    
#                 self.samples += 1
#         else:
#             job_position = 0
        
#         # Keep delay calculation
#         delay = self.calculate_delay(sqc_data, job_position) 
        
#         return job_position, delay


    def calculate_delay(self, sqc_data, job_position):  # Allowed delay
        current_pt = sqc_data[0]  # Current op time
        remaining_job_pt = sqc_data[1]  # Remaining time
        due_list = sqc_data[2]  # Due dates
        env_now = sqc_data[3]  # Current time
        remaining_no_op = sqc_data[10]  # Op index
        job_idx = sqc_data[-2][job_position]  # Job id
        
        t_curr = current_pt[job_position] if len(current_pt) > job_position else 0.0
        t_post = remaining_job_pt[job_position] if len(remaining_job_pt) > job_position else 0.0
        dd = due_list[job_position] if len(due_list) > job_position else 0.0 
        remaining_no_op_val = remaining_no_op[job_position] if len(remaining_no_op) > job_position else 0
        allowed_delay = dd - env_now - t_curr - t_post        
        opnum = self.m_no -  remaining_no_op_val
        idle_time_list = self.job_creator.idle_time[job_idx][1:opnum+1]
        # 3) Theoretical delay (alpha + conservative factor)
        if self.job_creator.production_record[job_idx][4] == True:
            if opnum <= 1:
                if allowed_delay > 0 :            
                    delay = allowed_delay / self.m_no
                else:
                    delay  = 0
            else:                
                if allowed_delay > 0  :
                    mean_time = np.mean(idle_time_list)
                    delay = max(0,min(mean_time,allowed_delay) - (env_now-self.job_creator.production_record[job_idx][0]))
                else:
                    delay  = 0
        else:
            if allowed_delay > 0 :
                d_ratio = 0;m_len=0;n_len=0
                for m in self.m_list:
                    if len(m.slack) > 0 :
                        m_len += len(m.slack) ; n_len += len([num for num in m.slack if num < 0])/len(m.slack)
                d_ratio = n_len / m_len     
                if d_ratio == 0:
                    d_ratio = random.random()   
                delay = allowed_delay * (1-d_ratio) 
                self.job_creator.idle_time[job_idx][0] = delay
            else:
                delay  = 0  
        
        return delay

    def compute_optimal_weights(self, A, k, w, beta=0.5, lambda_reg=0.1, min_weight=0.05, max_weight=0.95):
        """
        Optimize weights using Chebyshev score.
        Args:
            A: [n x 3] matrix of job objectives
            k: selected job index
            w: current weights
            lambda_reg: L2 regularization
            min_weight: min weight
            max_weight: max weight
        Returns:
            p: optimized weights
            new_score: new score for job k
        """
        n = A.shape[1]  # Objective count
        m = A.shape[0]  # Job count
        
        # Convert weights
        if torch.is_tensor(w):
            w_np = w.detach().cpu().numpy()
        else:
            w_np = np.array(w, dtype=np.float64)
        
        # Normalize objectives (consistent with selection_job_reward)
        min_vals = A.min(axis=0, keepdims=True)
        max_vals = A.max(axis=0, keepdims=True)
        ranges = max_vals - min_vals
        ranges[ranges < 1e-10] = 1.0
        A_normalized = (A - min_vals) / ranges
        
        # Chebyshev score
        def chebyshev_score(weights, preference):
            """Chebyshev score."""
            weighted = weights * preference
            return np.max(weighted)
        
        # Old score
        old_score = chebyshev_score(A_normalized[k], w_np)
        
        # Uniform weights for regularization
        uniform = np.ones(n) / n
        
        # Objective function
        def objective(p):
            # Key: minimize Chebyshev score for job k
            score_k = chebyshev_score(A_normalized[k], p)
            
            # L2 regularization
            reg = lambda_reg * np.sum((p - uniform)**2)
            
            return score_k + reg
        
        # Constraints
        constraints = []
        
        # 1) Sum to 1
        constraints.append({
            'type': 'eq',
            'fun': lambda p: np.sum(p) - 1
        })
        
        # 2) score_k <= score_j for all other jobs
        for j in range(m):
            if j != k:
                def make_constraint(j):
                    def constraint_func(p):
                        score_k = chebyshev_score(A_normalized[k], p)
                        score_j = chebyshev_score(A_normalized[j], p)
                        return score_j - score_k
                    return {'type': 'ineq', 'fun': constraint_func}
                constraints.append(make_constraint(j))
        
        # 3) Improvement constraint
        constraints.append({
            'type': 'ineq',
            'fun': lambda p: old_score - chebyshev_score(A_normalized[k], p)
        })
        
        # Bounds
        bounds = [(min_weight, max_weight) for _ in range(n)]
        
        # Initial value
        p0 = np.clip(w_np, min_weight, max_weight)
        p0 = p0 / np.sum(p0)
        
        # Solve
        try:
            res = minimize(
                objective,
                p0,
                bounds=bounds,
                constraints=constraints,
                method='SLSQP',
                options={'maxiter': 1000, 'ftol': 1e-8, 'disp': False}
            )
            
            if res.success:
                p_optimal = res.x
                p_optimal = np.clip(p_optimal, min_weight, max_weight)
                p_optimal = p_optimal / np.sum(p_optimal)
                
                # New score
                new_score = chebyshev_score(A_normalized[k], p_optimal)
                
                # Validate constraints
                constraints_ok = True
                tolerance = 1e-6
                
                # Ensure job k is not worse than others
                for j in range(m):
                    if j != k:
                        score_j = chebyshev_score(A_normalized[j], p_optimal)
                        if score_j < new_score - tolerance:
                            constraints_ok = False
                            break
                
                # Ensure improvement
                if new_score > old_score + tolerance:
                    constraints_ok = False
                
                if constraints_ok:
                    return torch.tensor(p_optimal, dtype=torch.float32), new_score
                else:
                    # Constraints not met
                    return torch.tensor(p0, dtype=torch.float32), old_score
            else:
                # Optimization failed
                return torch.tensor(p0, dtype=torch.float32), old_score
        
        except Exception as e:
            print(f"Optimization error: {e}")
            # Fallback to original
            return torch.tensor(p0, dtype=torch.float32), old_score

    def selection_job_reward(self, job_weights, current_preference):
        """
        Select a job via weighted Chebyshev score and compute reward.
        score_i = max_j { w_j * f_ij } (minimization, normalized)
        Args:
            job_weights: [n_jobs x 3] objectives (lower is better)
            current_preference: [3] weight vector
        Returns:
            selected_idx: index of selected job
            reward: minimum score
        """
        n_jobs = job_weights.shape[0]
        
        # Convert weights to numpy
        if torch.is_tensor(current_preference):
            w = current_preference.detach().cpu().numpy().flatten().astype(np.float32)
        else:
            w = np.array(current_preference, dtype=np.float32).flatten()
        
        # Normalize weights
        w_sum = np.sum(w)
        if w_sum > 0:
            w = w / w_sum
        else:
            # Use uniform weights if sum is zero
            w = np.ones_like(w) / len(w)
        
        # 1) Normalize objectives
        min_vals = job_weights.min(axis=0, keepdims=True)
        max_vals = job_weights.max(axis=0, keepdims=True)
        
        # Avoid division by zero
        ranges = max_vals - min_vals
        ranges[ranges < 1e-10] = 1.0
        
        # Normalize to [0, 1]
        normalized_weights = (job_weights - min_vals) / ranges
        
        # 2) Weighted Chebyshev score
        weighted = normalized_weights * w[np.newaxis, :]        
        scores = np.max(weighted, axis=1)
        
        # Key: pick the job with minimum score
        selected_idx = np.argmin(scores)
        reward = np.min(scores)
    
        return int(selected_idx), reward


class ContinuousSchedulingNetwork(nn.Module):
    def __init__(self, input_size,  preference_size=3):
        super(ContinuousSchedulingNetwork, self).__init__()
        
        # ========== Hyperparameters ==========
        self.lr = 0.0002
        self.input_size = input_size
        self.preference_size = preference_size
        self.clip_norm = 5.0
        
        # ========== Feature groups and normalization ==========
        # Used in forward
        self.base_info_size = 4
        self.global_info_size = 4  # base_info_size + 4 should be 8; kept per grouping
        self.pt_info_size = 3
        self.ttd_slack_info_size = 4
        self.heterogeneity_info_size = 3

        self.norm_base = nn.Sequential(nn.LayerNorm(4), nn.Flatten())
        self.norm_global = nn.Sequential(nn.LayerNorm(4), nn.Flatten())
        self.norm_pt = nn.Sequential(nn.LayerNorm(3), nn.Flatten())
        self.norm_ttd_slack = nn.Sequential(nn.LayerNorm(4), nn.Flatten())
        self.norm_heterogeneity = nn.Sequential(nn.LayerNorm(3), nn.Flatten())
        
        # ========== 1. State feature extractor ==========
        self.state_extractor = nn.Sequential(
            nn.Linear(18, 20),  # 18 -> 20
            nn.LeakyReLU(negative_slope=0.1),
            nn.LayerNorm(20),  # Match dimension
            nn.Dropout(0.1),
            nn.Linear(20, 16),  # Simplified
            nn.LeakyReLU(negative_slope=0.1),
            nn.LayerNorm(16),
            nn.Dropout(0.1),
            nn.Linear(16, 10),  # Output 10
            nn.LeakyReLU(negative_slope=0.1),
            nn.LayerNorm(10),
            nn.Dropout(0.1)
        )
        
        # ========== 2. Preference enhancer ==========
        self.pref_enhancer = nn.Sequential(
            nn.Linear(self.preference_size, 6),       # 3 -> 6
            nn.LeakyReLU(negative_slope=0.1),
            nn.LayerNorm(6),
            nn.Dropout(0.1),
            nn.Linear(6, self.preference_size),       # 6 -> 3
            nn.Softmax(dim=-1)
        )
        
        # ========== 3. Feature fusion ==========
        self.feature_fusion = nn.Sequential(
            nn.Linear(13, 10),
            nn.LeakyReLU(negative_slope=0.1),
            nn.LayerNorm(10),
            nn.Dropout(0.1),
            nn.Linear(10, 8),
            nn.LeakyReLU(negative_slope=0.1),
            nn.LayerNorm(8),
            nn.Dropout(0.1),
            nn.Linear(8, 6),
            nn.LeakyReLU(negative_slope=0.1)
        )
        self.residual_transform = nn.Linear(13, 6)  # Residual path
        
        # ========== 4. Dual heads ==========
        self.alpha_head = nn.Sequential(
            nn.Linear(6 + self.preference_size, 8),  # Fused + preference
            nn.LeakyReLU(negative_slope=0.1),
            nn.LayerNorm(8),
            nn.Dropout(0.1),
            nn.Linear(8, 4),
            nn.LeakyReLU(negative_slope=0.1),
            nn.LayerNorm(4),
            nn.Dropout(0.1),
            nn.Linear(4, 3),
            nn.Softmax(dim=-1)
        )
    
        # ========== 5. Direct preference path ==========
        self.pref_direct_alpha = nn.Sequential(
            nn.Linear(self.preference_size, 6),
            nn.LeakyReLU(negative_slope=0.1),
            nn.LayerNorm(6),
            nn.Dropout(0.1),
            nn.Linear(6, 3),
            nn.Softmax(dim=-1)
        )
        
        # ========== 6. Fusion gate ==========
        # Use gating instead of fixed weights
        self.fusion_gate = nn.Sequential(
            nn.Linear(6, 3),
            nn.LeakyReLU(negative_slope=0.1),
            nn.Linear(3, 1),
            nn.Sigmoid()
        )
        
        # ========== 7. Optimizer and scheduler ==========
        self.optimizer = optim.AdamW(
            self.parameters(),
            lr=self.lr,
            weight_decay=0.00005,
            betas=(0.9, 0.99),
            eps=1e-8,
            amsgrad=True
        )
        self.lr_scheduler = optim.lr_scheduler.ReduceLROnPlateau(
            self.optimizer,
            mode='min',
            factor=0.8,
            patience=1000,
            threshold=1e-4,
            min_lr=1e-6
        )
        
        # ========== Module list ==========
        self.network = nn.ModuleList([
            self.norm_base, self.norm_global, self.norm_pt,
            self.norm_ttd_slack, self.norm_heterogeneity,
            self.state_extractor, self.pref_enhancer,
            self.feature_fusion, self.residual_transform,
            self.alpha_head, self.pref_direct_alpha
        ])
        
        # ========== Init weights ==========
        self._init_weights()
            
    def _init_weights(self):
        """Initialize network weights."""
        for module in self.network:
            if isinstance(module, nn.Linear):
                nn.init.kaiming_normal_(module.weight, mode='fan_in', nonlinearity='leaky_relu', a=0.1)
                if module.bias is not None:
                    nn.init.constant_(module.bias, 0)
            elif isinstance(module, nn.LayerNorm):
                nn.init.constant_(module.weight, 1.0)
                nn.init.constant_(module.bias, 0)

    def forward(self, state_features, preference_vector):
        """
        Forward pass.
        Args:
            state_features: [batch_size, 18] raw state features
            preference_vector: [batch_size, 3] preference vector
        Returns:
            alpha_output: [batch_size, 3] scheduling weights
        """
        # 1) Normalize feature groups
        base_info = state_features[:, :4]
        global_info = state_features[:, 4:8]
        pt_info = state_features[:, 8:11]
        ttd_slack_info = state_features[:, 11:15]
        heterogeneity_info = state_features[:, 15:18]
        
        # Apply normalization
        base_norm = self.norm_base(base_info)
        global_norm = self.norm_global(global_info)
        pt_norm = self.norm_pt(pt_info)
        ttd_slack_norm = self.norm_ttd_slack(ttd_slack_info)
        heterogeneity_norm = self.norm_heterogeneity(heterogeneity_info)
        
        # Concatenate normalized features
        normalized_features = torch.cat([
            base_norm, global_norm, pt_norm, 
            ttd_slack_norm, heterogeneity_norm
        ], dim=-1)
        
        # 2) State feature extraction
        state_features = self.state_extractor(normalized_features)
        
        # 3) Preference enhancement
        enhanced_pref = self.pref_enhancer(preference_vector)
        
        # 4) Feature fusion with residual
        combined_features = torch.cat([state_features, enhanced_pref], dim=-1)
        fusion_out = self.feature_fusion(combined_features)
        residual = self.residual_transform(combined_features)
        
        # Residual connection
        fused_features = fusion_out + residual
        
        # 5) Dual-path output
        # Path A: fused + preference
        alpha_from_fusion = self.alpha_head(
            torch.cat([fused_features, enhanced_pref], dim=-1)
        )
        
        # Path B: direct preference
        alpha_from_pref = self.pref_direct_alpha(enhanced_pref)
        
        # 6) Adaptive fusion (gated)
        gate = self.fusion_gate(fused_features)
        alpha_output = gate * alpha_from_fusion + (1 - gate) * alpha_from_pref
        
        return alpha_output


class DiscreteSequencingBrain:
    """Shared base class for discrete sequencing."""
    
    def __init__(self, env, job_creator, all_machines, job_numbers, *args, **kwargs):
        self.env = env  # SimPy environment
        self.job_creator = job_creator  # Job generator
        self.m_list = all_machines  # Machine list
        self.m_no = len(self.m_list)  # Machine count
        self.warm_up = 0.01 * job_numbers * job_creator.avg_pt  # Warm-up time
        self.span = job_numbers * job_creator.avg_pt  # Total sim time
        self.job_creator.build_sqc_experience_repository(self.m_list)  # Replay buffer
        
        # Common parameters
        self.minibatch_size = 128  # Batch size
        self.sequencing_action_NN_training_interval = 50  # Training interval
        self.sequencing_action_NN_training_time_record = []  # Training time record
        self.discount_factor = 0.99  # Discount factor
        self.epsilon = 0.1  # Exploration rate
        self.loss_time_record = []  # Loss time record
        self.loss_record = []  # Loss record
        self.rule_loss_record = []  # Rule loss record
        self.alpha_loss_record = []  # Alpha loss record
        self.sample_reward_record = []  # Reward record
        self.train_reward_record = []        
        self.current_preference_vector = kwargs.get('preference_vector', None)
        self.train = kwargs.get('train', False)    
        self.address = kwargs.get('address', None)
        self.ablation = kwargs.get('ablation', None)
        self.lambda_param = 0.7  # TD(lambda)
        self.baseline_ema_alpha = 0.99  # Baseline EMA
        self.running_baseline = 0.0  # Running baseline
        self.mixed_training_ratio = 0.7  # PG vs supervised ratio
        self.trajectory_buffer_size = 1280  # Min trajectory length
        self.good_experience_threshold = 0.6  # Good experience threshold
        self.samples = 0
        
        # DQN parameters
        self.max_queue_size = 5  # Max queue size for outputs
        self.target_update_freq = 100  # Target update frequency
        self.tau = 0.01  # Soft update
        
        # Experience buffer
        self.trajectory_buffer = []  # Trajectory buffer
        
        # Initialize sequencing methods
        for m in self.m_list:           
            m.job_sequencing = self.action_default  # Default sequencing
        
        self.func_list = [sequencing.CR, sequencing.MDD, sequencing.MOD, 
                         sequencing.MS, sequencing.ATC, sequencing.EDD]  # Candidate rules
        
        self.multi_obj_manager = mutilobjectivemanager.MultiObjectiveManager(num_objectives=3)           
        self.training_step_count = 0
        self.sampling_times = 0
        
        # Implement in subclasses
        self.sequencing_action_NN = None
        self.target_network = None
        self.input_size = None
        
    def _init_networks(self):
        """Initialize networks (subclass)."""
        raise NotImplementedError
        
    def warm_up_process(self):  # Warm-up
        for m in self.m_list:
             m.job_sequencing = self.action_warm_up
        for idx, func in enumerate(self.func_list):            
            self.func_selection = idx
            print('Time {}: rule set to {}'.format(self.env.now, func))
            yield self.env.timeout(int(self.warm_up / (len(self.func_list) + 1)))  
            
        for m in self.m_list:
            m.job_sequencing = self.action_default 
            
        print("Time {} to {}: random rules for all machines.".format(self.env.now, self.warm_up))
        yield self.env.timeout(self.warm_up - self.env.now - 1)        
        
        print("Time {}: machines start RL-based sequencing.".format(self.env.now))
        for m in self.m_list:
            m.job_sequencing = self.action_sqc_rule 

    def update_learning(self):  # Learning-rate update
        yield self.env.timeout(self.warm_up)
        initial_lr = self.sequencing_action_NN.optimizer.param_groups[0]['lr']
        target_lr = initial_lr / 10
        reduction = (initial_lr - target_lr) / 10
        lr_min = 1e-6  # Avoid negative learning rate
        while self.env.now < self.span:
            yield self.env.timeout((self.span - self.warm_up) / 10)
            new_lr = self.sequencing_action_NN.optimizer.param_groups[0]['lr'] - reduction
            self.sequencing_action_NN.optimizer.param_groups[0]['lr'] = max(new_lr, lr_min)
            
    # Default action phase
    def action_default(self, sqc_data):
        current_queue = sqc_data[-2]
        queue_size = len(current_queue)
        m_idx = sqc_data[-1]
        if queue_size > 1:
            s_t = state_multi_channel(self, sqc_data)
            job_weights = self.multi_obj_manager.calculate_importance(self.job_creator, current_queue)
            preference = self.multi_obj_manager.calculate_preference(max(current_queue))
            job_probs = torch.tensor([random.random() for _ in range(self.max_queue_size)], dtype=torch.float32)
            job_position = random.randint(0, queue_size - 1)
            if job_position < self.max_queue_size:
                job_probs[job_position] = 0   
            job_weights_tensor = torch.tensor(job_weights, dtype=torch.float32)
            job_probs_tensor = torch.tensor(job_probs, dtype=torch.float32)   
            _, reward = self.selection_job_reward(job_weights_tensor, job_probs_tensor)
            if self.train == True:    
                build_experience(self, self.env.now, m_idx, s_t, job_probs_tensor, reward, preference, job_weights)           
                
        else:
            job_position = 0
        delay = self.calculate_delay(sqc_data, job_position) if not (self.ablation == "Ablation0" or self.ablation == "Ablation2") else 0            
            
        return job_position, delay 
      
    # Warm-up action phase
    def action_warm_up(self, sqc_data):       
        current_queue = sqc_data[-2]
        queue_size = len(current_queue)
        m_idx = sqc_data[-1]
        if queue_size > 1:
            s_t = state_multi_channel(self, sqc_data) 
            job_position, _ = self.func_list[self.func_selection](sqc_data)            
            job_weights = self.multi_obj_manager.calculate_importance(self.job_creator, current_queue)
            preference = self.multi_obj_manager.calculate_preference(max(current_queue)) 
            job_probs = torch.tensor([random.random() for _ in range(self.max_queue_size)], dtype=torch.float32)
            job_position = random.randint(0, queue_size - 1)
            if job_position < self.max_queue_size:
                job_probs[job_position] = 0 
            job_weights_tensor = torch.tensor(job_weights, dtype=torch.float32)
            job_probs_tensor = torch.tensor(job_probs, dtype=torch.float32)
            _, reward = self.selection_job_reward(job_weights_tensor, job_probs_tensor)
            if self.train == True:               
                build_experience(self, self.env.now, m_idx, s_t, job_probs_tensor, reward, preference, job_weights)           
        else:
            job_position = 0
        delay = self.calculate_delay( sqc_data, job_position) if not (self.ablation == "Ablation0" or self.ablation == "Ablation2") else 0             
        return job_position, delay  

    # Random exploration phase
    def action_ablation(self, sqc_data):                         
        current_queue = sqc_data[-2]
        queue_size = len(current_queue)    
        if queue_size > 1:  
            order_coeff = random.random()
            job_position = int(round(order_coeff * (queue_size - 1)))
            job_position = max(0, min(job_position, queue_size - 1))
        else:
            job_position = 0
        delay = self.calculate_delay( sqc_data, job_position) if not (self.ablation == "Ablation0" or self.ablation == "Ablation2") else 0            
        return job_position, delay

        def calculate_delay(self, sqc_data, job_position):  # Allowed delay
            current_pt = sqc_data[0]  # Current op time
            remaining_job_pt = sqc_data[1]  # Remaining time
            due_list = sqc_data[2]  # Due dates
            env_now = sqc_data[3]  # Current time
            remaining_no_op = sqc_data[10]  # Op index
            job_idx = sqc_data[-2][job_position]  # Job id
            
            t_curr = current_pt[job_position] if len(current_pt) > job_position else 0.0
            t_post = remaining_job_pt[job_position] if len(remaining_job_pt) > job_position else 0.0
            dd = due_list[job_position] if len(due_list) > job_position else 0.0 
            remaining_no_op_val = remaining_no_op[job_position] if len(remaining_no_op) > job_position else 0
            allowed_delay = dd - env_now - t_curr - t_post        
            opnum = self.m_no -  remaining_no_op_val
            idle_time_list = self.job_creator.idle_time[job_idx][1:opnum+1]
            # 3) Theoretical delay (alpha + conservative factor)
            if self.job_creator.production_record[job_idx][4] == True:
                if opnum <= 1:
                    if allowed_delay > 0 :            
                        delay = allowed_delay / self.m_no
                    else:
                        delay  = 0
                else:                
                    if allowed_delay > 0  :
                        mean_time = np.mean(idle_time_list)
                        delay = max(0,min(mean_time,allowed_delay) - (env_now-self.job_creator.production_record[job_idx][0]))
                    else:
                        delay  = 0
            else:
                if allowed_delay > 0 :
                    d_ratio = 0;m_len=0;n_len=0
                    for m in self.m_list:
                        if len(m.slack) > 0 :
                            m_len += len(m.slack) ; n_len += len([num for num in m.slack if num < 0])/len(m.slack)
                    d_ratio = n_len / m_len     
                    if d_ratio == 0:
                        d_ratio = random.random()   
                    delay = allowed_delay * (1-d_ratio) 
                    self.job_creator.idle_time[job_idx][0] = delay
                else:
                        delay  = 0  
            
            return delay

    # # RL action phase (base version)
    # def action_sqc_rule(self, sqc_data):
    #     current_queue = sqc_data[-2]   
    #     queue_size = len(current_queue)
    #     m_idx = sqc_data[-1]
    #     if queue_size > 1:
    #         s_t = state_multi_channel(self, sqc_data) 
    #         s_t_reshaped = s_t.reshape([1, self.input_size]) 
    #         if self.train == False:  # Inference
    #             preference = self.current_preference_vector                
    #         else:  # Training
    #             preference = self.multi_obj_manager.calculate_preference(max(current_queue))
    #         job_weights = self.multi_obj_manager.calculate_importance(self.job_creator, current_queue)
               
    #         if self.ablation == "Ablation4":
    #             preference = torch.tensor(np.array(np.random.dirichlet([1, 1, 1])), dtype=torch.float32)
                
    #         # Exploration strategy (subclass)
    #         job_probs = self._get_action_probabilities(s_t_reshaped, preference, queue_size)
            
    #         job_weights_tensor = torch.tensor(job_weights, dtype=torch.float32) 
    #         job_probs_tensor = torch.tensor(job_probs, dtype=torch.float32)          
    #         job_position, reward = self.selection_job_reward(job_weights_tensor, job_probs_tensor)
            
    #         if self.train == True:                              
    #             build_experience(self, self.env.now, m_idx, s_t, job_probs_tensor, reward, preference, job_weights)           
    #             self.samples += 1
    #     else:
    #         job_position = 0     
             
    #     delay = self.calculate_delay( sqc_data, job_position) if not (self.ablation == "Ablation0" or self.ablation == "Ablation2") else 0             
        
    #     return job_position, delay
    
    def action_sqc_rule(self, sqc_data):
        """Integrated PD-MORL (three components)."""
        current_queue = sqc_data[-2]
        queue_size = len(current_queue)
        m_idx = sqc_data[-1]
        
        if queue_size > 1:
            s_t = state_multi_channel(self, sqc_data)
            s_t_reshaped = s_t.reshape([1, self.input_size])
            
            # Job importance (original reward)
            job_weights = self.multi_obj_manager.calculate_importance(self.job_creator, current_queue)
            
            if self.train == False:  # Inference
                # Keep original logic
                preference = self.current_preference_vector 
            else:  # Training
                # ========== PD-MORL 1: preference-driven update ==========
                preference = self.pd_morl_manager.sample_preference()    
            
            # Action selection
            if random.random() < self.epsilon :
                order_coeff = torch.tensor(np.random.dirichlet([1, 1, 1]), dtype=torch.float32)
            else:
                # ========== PD-MORL: network takes (state, preference) ==========
                order_coeff = self.network.forward(s_t_reshaped, preference.reshape(1, 3))
                order_coeff = order_coeff.squeeze(0)
            
            # ========== Original reward kept ==========
            job_position, original_reward = self.selection_job_reward(job_weights, order_coeff)
            
            # ========== PD-MORL 2: cosine similarity term ==========
            if self.train:
                # Cosine similarity
                cosine_sim = F.cosine_similarity(
                    order_coeff.unsqueeze(0),
                    preference.unsqueeze(0)
                ).item()
                # HER preference samples
                her_preferences = self.pd_morl_manager.get_her_preferences(preference, job_weights[job_position, :])
                
                # Store experience per preference
                for her_pref in her_preferences:
                    # Recompute reward for preference
                    _, her_reward = self.selection_job_reward(job_weights, her_pref)
                    if torch.allclose(her_pref, preference):
                        final_reward = her_reward * (1.0 + self.cosine_weight * max(0, cosine_sim))
                    else:
                        final_reward = her_reward
                        
                build_experience(self, self.env.now, m_idx, s_t, order_coeff.detach(), final_reward, preference, job_weights)                    
                self.samples += 1
        else:
            job_position = 0
        
        # Keep delay calculation
        delay = self.calculate_delay(sqc_data, job_position) 
        
        return job_position, delay


    def _get_action_probabilities(self, state, preference, queue_size):
        """Get action probabilities (subclass)."""
        raise NotImplementedError
  

    def selection_job_reward(self, job_weights, job_probabilities):        
        if isinstance(job_weights, np.ndarray):
            job_weights = torch.from_numpy(job_weights).float()

        n_jobs = job_weights.shape[0]
        max_output_size = len(job_probabilities)          
        # Preprocess
        if n_jobs > max_output_size:
            # Only consider first max_output_size jobs
            job_weights = job_weights[:max_output_size, :]
            job_probs = job_probabilities  # Use all probabilities
            actual_n = max_output_size
        else:
            # Queue size <= output size
            job_weights = job_weights  # Use all jobs
            job_probs = job_probabilities[:n_jobs]  # Use first n probabilities
            actual_n = n_jobs        
        # Normalize weights (optional)
        epsilon = 1e-8      
        normalized_weights = job_weights / (torch.sum(job_weights, dim=0, keepdim=True) + epsilon)
        # Compute scores
        scores = torch.zeros(actual_n, device=job_probabilities.device)
        beta = 0.5  # Weight for sum vs product
        for i in range(actual_n):
            f = normalized_weights[i, :]
            f_safe = torch.maximum(f, torch.tensor(epsilon, device=f.device))
            job_prob = torch.maximum(job_probs[i], torch.tensor(epsilon, device=job_probs.device))
            weighted_sum = job_prob * torch.sum(f_safe)
            log_sum = job_prob * torch.sum(torch.log(f_safe))
            product_term = torch.exp(log_sum) if log_sum != 0 else torch.tensor(0.0, device=f.device)
            total_score = beta * weighted_sum + (1 - beta) * product_term
            scores[i] = total_score
    
        # Select min score
        selected_idx = torch.argmin(scores)
        reward = torch.min(scores)    
        
        return selected_idx, reward
    
    def training_process_parameter_sharing(self): 
        yield self.env.timeout(self.warm_up + 1)
        # After warm-up, supervised learning on warm-up data
        if len(self.trajectory_buffer) > self.minibatch_size:
            for i in range(10):  # Supervised steps
                self._train_from_replay()
        
        # Training
        while self.job_creator.in_system_job_no >= 1:
            if self.samples >= self.minibatch_size:
                self._train_from_replay(num_training_rounds=5)
                self.samples = self.samples - self.minibatch_size
                # Periodic target update
            if self.training_step_count % self.target_update_freq == 0:
                self.update_target_network()
            yield self.env.timeout(100)
        
        # Save model
        address = self.address.format(sys.path[0])
        torch.save(self.sequencing_action_NN.state_dict(), address)
        pref_address = address.replace('.pt', '_preferences.pkl')
        self.multi_obj_manager.save_training_results(pref_address)
        print(f"Model and preferences saved: {address}, {pref_address}")
    
    def _train_from_replay(self, num_training_rounds=5):
        """Train from replay (subclass)."""
        raise NotImplementedError
    
    def update_target_network(self):
        """Soft-update target network."""
        for target_param, param in zip(self.target_network.parameters(), self.sequencing_action_NN.parameters()):
            target_param.data.copy_(self.tau * param.data + (1.0 - self.tau) * target_param.data)


class DiscreteSchedulingNetwork(nn.Module):
    """Shared base scheduling network."""
    
    def __init__(self, input_size, max_queue_size=5, preference_size=3):
        super(DiscreteSchedulingNetwork, self).__init__()
        
        # ========== Hyperparameters ==========
        self.lr = 0.0002
        self.input_size = input_size
        self.max_queue_size = max_queue_size  # Output size
        self.preference_size = preference_size
        self.clip_norm = 5.0
        
        # ========== Feature groups and normalization ==========
        self.base_info_size = 4
        self.global_info_size = 4
        self.pt_info_size = 3
        self.ttd_slack_info_size = 4
        self.heterogeneity_info_size = 3

        self.norm_base = nn.Sequential(nn.LayerNorm(4), nn.Flatten())
        self.norm_global = nn.Sequential(nn.LayerNorm(4), nn.Flatten())
        self.norm_pt = nn.Sequential(nn.LayerNorm(3), nn.Flatten())
        self.norm_ttd_slack = nn.Sequential(nn.LayerNorm(4), nn.Flatten())
        self.norm_heterogeneity = nn.Sequential(nn.LayerNorm(3), nn.Flatten())
        
        # Implement in subclass
        self._init_network_layers()
        
        # ========== Optimizer and scheduler ==========
        self.optimizer = optim.AdamW(
            self.parameters(),
            lr=self.lr,
            weight_decay=0.00005,
            betas=(0.9, 0.99),
            eps=1e-8,
            amsgrad=True
        )
        self.lr_scheduler = optim.lr_scheduler.ReduceLROnPlateau(
            self.optimizer,
            mode='min',
            factor=0.8,
            patience=1000,
            threshold=1e-4,
            min_lr=1e-6
        )
        
        # ========== Init weights ==========
        self._init_weights()
    
    def _init_network_layers(self):
        """Initialize network layers (subclass)."""
        raise NotImplementedError
            
    def _init_weights(self):
        """Initialize network weights."""
        for module in self.modules():
            if isinstance(module, nn.Linear):
                nn.init.kaiming_normal_(module.weight, mode='fan_in', nonlinearity='leaky_relu', a=0.1)
                if module.bias is not None:
                    nn.init.constant_(module.bias, 0)
            elif isinstance(module, nn.LayerNorm):
                nn.init.constant_(module.weight, 1.0)
                nn.init.constant_(module.bias, 0)

    def forward(self, state_features, preference_vector):
        """Forward pass (subclass)."""
        raise NotImplementedError