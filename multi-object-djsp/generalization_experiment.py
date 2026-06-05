# main_training_GC.py
import simpy
import sys
import numpy as np
import datetime
import os
import pickle
import torch
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
import importlib
from common.experiment_scene import orthogonal_scenarios, find_scenario
from common.shared_modules import (ShopFloor, MultiObjectiveManager)

plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans', 'Arial Unicode MS', 'Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False
plt.rcParams['figure.dpi'] = 100
plt.rcParams['savefig.dpi'] = 300
# ============================================
# 1. Experiment config
# ============================================

# Algorithm selection (PG-HRL only)
ALGORITHM = 'PG_HRL'
brain_machine = importlib.import_module(f"algorithm.RL.brain_{ALGORITHM}") 
# Scenario group definitions
SCENARIO_GROUPS = {
    'A': ['L9-2', 'L9-5', 'L9-6'],  # Mild balanced group
    'B': ['L9-1', 'L9-3', 'L9-8', 'L9-9'],  # Single-stress group
    'C': ['L9-4', 'L9-7']  # Multi-stress group
}

# Representative training models (one per group)
REPRESENTATIVE_MODELS = {
    'A': 'L9-6',  # Mid-scale stable production
    'B': 'L9-8',  # Large-scale high load
    'C': 'L9-4'   # Mid-scale high pressure
}

# All test scenarios (9 orthogonal scenarios)
ALL_TEST_SCENARIOS = [scenario['scenario_id'] for scenario in orthogonal_scenarios]

# Experiment parameters
EXPERIMENT_CONFIG = {
    'test_repetitions': 10,
    'preference_perturbations': 5,
    'simulation_span': 50,
    'noise_level': 0.08,
    'random_seed_range': 20000
}

# ============================================
# 2. Helper classes and functions
# ============================================

class SceneGeneralizationAnalyzer:
    """Scene generalization analyzer (fixed scene, compare models)."""
    
    @staticmethod
    def analyze_scene_generalization(test_scenario_id, all_model_results):
        """
        Analyze generalization across models for a single test scene.
        Steps:
        1) Collect all model solutions for the test scene
        2) Compute unified non-dominated set
        3) Split non-dominated solutions by model
        4) Compute per-model metrics
        5) Analyze adaptability
        """
        print(f"\n{'='*60}")
        print(f"Scene generalization analysis - test scene: {test_scenario_id}")
        print(f"{'='*60}")
        
        # 1. Collect all model solutions for the test scene
        all_model_solutions = {}
        total_solutions_count = 0
        
        for train_scenario_id, model_results in all_model_results.items():
            if test_scenario_id in model_results and model_results[test_scenario_id]:
                result = model_results[test_scenario_id]
                
                # Extract solutions from raw data
                solutions_for_model = SceneGeneralizationAnalyzer._extract_solutions_from_raw_data(
                    result['raw_data']
                )
                
                all_model_solutions[train_scenario_id] = solutions_for_model
                total_solutions_count += len(solutions_for_model)
                print(f"  Model {train_scenario_id}: {len(solutions_for_model)} solutions")
        
        if len(all_model_solutions) < 2:
            print("Warning: need at least 2 models for comparison")
            return None
        
        # 2. Compute unified non-dominated set across all models for this test scene
        all_solutions = []
        solution_source_info = []  # to track which model each solution came from
        
        for model_id, solutions in all_model_solutions.items():
            for i, sol in enumerate(solutions):
                all_solutions.append(sol)
                solution_source_info.append({
                    'train_model': model_id,
                    'solution_index': i,
                    'objectives': sol
                })
        
        print(f"Scene {test_scenario_id} total solutions: {len(all_solutions)} (from {len(all_model_solutions)} models)")
        
        # 3. Compute unified non-dominated set
        print("\nComputing unified non-dominated set for this scene...")
        solutions_array = np.array(all_solutions)
        
        # Identify non-dominated solutions
        n = len(solutions_array)
        is_pareto = np.ones(n, dtype=bool)
        
        for i in range(n):
            if not is_pareto[i]:
                continue
            for j in range(n):
                if i != j and is_pareto[j]:
                    # Check if solution j dominates solution i
                    if np.all(solutions_array[j] <= solutions_array[i]) and np.any(solutions_array[j] < solutions_array[i]):
                        is_pareto[i] = False
                        break
        
        # Extract non-dominated solutions and their source info
        pareto_solutions = solutions_array[is_pareto]
        pareto_info = [solution_source_info[i] for i in range(n) if is_pareto[i]]
        
        print(f"Non-dominated solutions: {len(pareto_solutions)} / {n} ({len(pareto_solutions)/n*100:.1f}%)")
        
        # 4. Split non-dominated solutions by training model
        print("\nSplitting non-dominated solutions by training model...")
        model_pareto_contributions = {}
        
        for info in pareto_info:
            model_id = info['train_model']
            if model_id not in model_pareto_contributions:
                model_pareto_contributions[model_id] = []
            model_pareto_contributions[model_id].append(info['objectives'])
        
        # 5. Compute per-model metrics
        print("\nComputing per-model generalization metrics...")
        model_metrics = {}
        
        # Temporary MultiObjectiveManager for normalization
        temp_mo_manager = MultiObjectiveManager()
        
        # Add all solutions to compute global reference points
        for i, sol in enumerate(all_solutions):
            temp_mo_manager.add_experiment_data(
                'ALL_MODELS',
                f'solution_{i}',
                [{'objectives': sol}]
            )
        
        # Global reference points
        temp_mo_manager.calculate_global_reference_points()
        
        # Extract normalization params
        global_ideal = temp_mo_manager.global_ideal
        normalization_ranges = temp_mo_manager.normalization_ranges
        
        for model_id in all_model_solutions.keys():
            if model_id in model_pareto_contributions:
                # This model's contribution on the non-dominated set
                model_pareto_solutions = model_pareto_contributions[model_id]
                model_pareto_count = len(model_pareto_solutions)
                
                # Total solutions for this model
                model_total_solutions = len(all_model_solutions[model_id])
                
                if model_pareto_count > 0:
                    # Coverage ratio (pareto contribution)
                    coverage_ratio = model_pareto_count / len(pareto_solutions)
                    
                    # Success ratio (solutions on pareto front)
                    success_ratio = model_pareto_count / model_total_solutions
                    
                    # Hypervolume (normalized)
                    normalized_solutions = (np.array(model_pareto_solutions) - global_ideal) / normalization_ranges
                    normalized_solutions = np.clip(normalized_solutions, 0, 1)
                    
                    hv = SceneGeneralizationAnalyzer._calculate_hypervolume_normalized(normalized_solutions)
                    
                    # Spacing
                    spacing = SceneGeneralizationAnalyzer._calculate_spacing_metric_normalized(normalized_solutions)
                    
                    # Convergence
                    convergence = SceneGeneralizationAnalyzer._calculate_convergence_to_ideal_normalized(normalized_solutions)
                    
                    # Generalization index
                    model_generalization_index = SceneGeneralizationAnalyzer._calculate_model_generalization_index(
                        hv, coverage_ratio, success_ratio, spacing, convergence
                    )
                    
                    model_metrics[model_id] = {
                        'train_scenario': model_id,
                        'solution_count': model_total_solutions,
                        'pareto_count': model_pareto_count,
                        'coverage_ratio': coverage_ratio,
                        'success_ratio': success_ratio,
                        'hypervolume': hv,
                        'spacing': spacing,
                        'convergence': convergence,
                        'generalization_index': model_generalization_index,
                        'generalization_level': SceneGeneralizationAnalyzer._get_generalization_level(model_generalization_index)
                    }
                    
                    print(f"  Model {model_id}: "
                          f"pareto={model_pareto_count}, "
                          f"coverage={coverage_ratio:.2%}, "
                          f"success={success_ratio:.2%}, "
                          f"HV={hv:.4f}, "
                          f"GI={model_generalization_index:.4f}")
                else:
                      # No pareto contributions
                    model_metrics[model_id] = {
                        'train_scenario': model_id,
                        'solution_count': model_total_solutions,
                        'pareto_count': 0,
                        'coverage_ratio': 0,
                        'success_ratio': 0,
                        'hypervolume': 0,
                        'spacing': 5.0,
                        'convergence': 5.0,
                        'generalization_index': 0,
                        'generalization_level': "Poor"
                    }
                    print(f"  Model {model_id}: no pareto contributions")
        
        # 6) Scene-level comparison metrics
        print("\nComputing scene-level comparison metrics...")
        scene_comparison = SceneGeneralizationAnalyzer._calculate_scene_comparison_metrics(
            model_metrics, len(pareto_solutions), total_solutions_count
        )
        
        # 7) Group-level analysis
        group_analysis = SceneGeneralizationAnalyzer._analyze_group_performance(
            test_scenario_id, model_metrics
        )
        
        # Best model for the scene
        best_model = scene_comparison['best_model']
        if best_model and best_model in model_metrics:
            best_gi = model_metrics[best_model]['generalization_index']
            print(f"\nScene {test_scenario_id} best model: {best_model} (GI={best_gi:.4f})")
        
        return {
            'test_scenario': test_scenario_id,
            'total_solutions': total_solutions_count,
            'total_pareto_solutions': len(pareto_solutions),
            'pareto_ratio': len(pareto_solutions) / total_solutions_count if total_solutions_count > 0 else 0,
            'model_metrics': model_metrics,
            'scene_comparison': scene_comparison,
            'group_analysis': group_analysis,
            'global_ideal': global_ideal,
            'global_nadir': temp_mo_manager.global_nadir
        }
    
    @staticmethod
    def _extract_solutions_from_raw_data(raw_data):
        """Extract objective vectors from raw data."""
        solutions = []
        for rule_data in raw_data.values():
            for run_data in rule_data.values():
                for job_data in run_data:
                    if 'objectives' in job_data:
                        solutions.append(job_data['objectives'])
        return solutions
    
    @staticmethod
    def _calculate_hypervolume_normalized(normalized_solutions, ref_point=None):
        """Compute hypervolume on normalized solutions."""
        if len(normalized_solutions) <= 1:
            return 0
        
        if ref_point is None:
            ref_point = np.ones(normalized_solutions.shape[1]) * 1.1
        
        try:
            from pymoo.indicators.hv import HV
            ind = HV(ref_point=ref_point)
            hv_value = ind(normalized_solutions)
            return hv_value
        except ImportError:
            # Simplified fallback
            n_solutions = len(normalized_solutions)
            if n_solutions == 0:
                return 0
            
            dominated_volume = 1.0
            for dim in range(normalized_solutions.shape[1]):
                min_val = np.min(normalized_solutions[:, dim])
                dominated_volume *= (ref_point[dim] - min_val)
            
            return dominated_volume / n_solutions
    
    @staticmethod
    def _calculate_spacing_metric_normalized(normalized_solutions):
        """Compute spacing metric on normalized solutions."""
        n = len(normalized_solutions)
        if n <= 1:
            return 5.0
        
        distances = []
        for i in range(n):
            min_dist = float('inf')
            for j in range(n):
                if i != j:
                    dist = np.linalg.norm(normalized_solutions[i] - normalized_solutions[j])
                    if dist < min_dist:
                        min_dist = dist
            if min_dist < float('inf'):
                distances.append(min_dist)
        
        if len(distances) == 0:
            return 5.0
        
        mean_dist = np.mean(distances)
        if mean_dist == 0:
            return 0
        
        std_dist = np.std(distances)
        return std_dist / mean_dist
    
    @staticmethod
    def _calculate_convergence_to_ideal_normalized(normalized_solutions):
        """Compute convergence to ideal point (normalized)."""
        n = len(normalized_solutions)
        if n <= 1:
            return 5.0
        
        ideal_point = np.zeros(normalized_solutions.shape[1])
        distances = np.linalg.norm(normalized_solutions - ideal_point, axis=1)
        return np.mean(distances)
    
    @staticmethod
    def _calculate_model_generalization_index(hv, coverage_ratio, success_ratio, spacing, convergence):
        """Compute model generalization index (standard)."""
        
        # Hypervolume score
        if hv >= 1.0:
            hv_score = 1.0 + min(0.15, (hv - 1.0))
        else:
            hv_score = hv
        
        # Coverage score
        coverage_score = coverage_ratio
        
        # Key: weighted GI
        gi = hv_score * 0.7 + coverage_score * 0.3
        
        # Clamp to range
        return min(1.2, max(0.0, gi))
    
    @staticmethod
    def _get_generalization_level(gi):
        """Map GI to generalization level."""
        if gi >= 1.0:
            return "Excellent"
        elif gi >= 0.8:
            return "Very good"
        elif gi >= 0.6:
            return "Good"
        elif gi >= 0.4:
            return "Fair"
        else:
            return "Poor"
    
    @staticmethod
    def _calculate_scene_comparison_metrics(model_metrics, total_pareto_count, total_solutions_count):
        """Compute scene-level comparison metrics."""
        if not model_metrics:
            return {
                'best_model': None,
                'worst_model': None,
                'average_gi': 0,
                'gi_std': 0,
                'performance_gap': 0,
                'dominance_ratio': 0
            }
        
        # Identify the best and worst models
        models = list(model_metrics.keys())
        gis = [metrics['generalization_index'] for metrics in model_metrics.values()]
        
        best_idx = np.argmax(gis)
        worst_idx = np.argmin(gis)
        
        best_model = models[best_idx]
        worst_model = models[worst_idx]
        
        # Compute dominance relationships
        dominance_counts = {}
        for model_id in models:
            dominance_counts[model_id] = 0
        
        for model_id, metrics in model_metrics.items():
            for other_id, other_metrics in model_metrics.items():
                if model_id != other_id:
                    if metrics['pareto_count'] > other_metrics['pareto_count']:
                        dominance_counts[model_id] += 1
        
        # Compute average dominance ratio
        dominance_ratio = np.mean([count/(len(models)-1) for count in dominance_counts.values()]) if len(models) > 1 else 0
        
        return {
            'best_model': best_model,
            'worst_model': worst_model,
            'average_gi': np.mean(gis),
            'gi_std': np.std(gis),
            'performance_gap': gis[best_idx] - gis[worst_idx],
            'dominance_ratio': dominance_ratio,
            'total_pareto_solutions': total_pareto_count,
            'total_solutions': total_solutions_count
        }
    
    @staticmethod
    def _analyze_group_performance(test_scenario_id, model_metrics):
        """Analyze group differences (train vs. test group relation)."""
        # Determine the test scenario group
        test_group = None
        for group, scenarios in SCENARIO_GROUPS.items():
            if test_scenario_id in scenarios:
                test_group = group
                break
        
        if test_group is None:
            return None
        
        # Group statistics
        same_group_gis = []
        different_group_gis = []
        
        for model_id, metrics in model_metrics.items():
            # Determine the training scenario group
            train_group = None
            for group, scenarios in SCENARIO_GROUPS.items():
                if model_id in scenarios:
                    train_group = group
                    break
            
            if train_group is None:
                continue
            
            gi = metrics['generalization_index']
            
            if train_group == test_group:
                same_group_gis.append(gi)
            else:
                different_group_gis.append(gi)
        
        return {
            'test_scenario_group': test_group,
            'same_group_avg': np.mean(same_group_gis) if same_group_gis else 0,
            'same_group_std': np.std(same_group_gis) if same_group_gis else 0,
            'different_group_avg': np.mean(different_group_gis) if different_group_gis else 0,
            'different_group_std': np.std(different_group_gis) if different_group_gis else 0,
            'generalization_advantage': (np.mean(same_group_gis) if same_group_gis else 0) - 
                                       (np.mean(different_group_gis) if different_group_gis else 0)
        }


class CrossSceneAnalyzer:
    """Cross-scene analyzer."""
    
    @staticmethod
    def analyze_overall_generalization(all_scene_analysis):
        """Aggregate scenes to analyze overall model generalization."""
        print(f"\n{'='*60}")
        print("Cross-scene generalization analysis")
        print(f"{'='*60}")
        
        if not all_scene_analysis:
            print("No data available for analysis")
            return None
        
        # Collect per-model performance across scenes
        model_performance = {}
        
        for test_scenario, analysis in all_scene_analysis.items():
            if analysis is None:
                continue
            
            model_metrics = analysis['model_metrics']
            
            for model_id, metrics in model_metrics.items():
                if model_id not in model_performance:
                    model_performance[model_id] = {
                        'gi_scores': [],
                        'coverage_ratios': [],
                        'success_ratios': [],
                        'hypervolumes': [],
                        'best_scenes': [],  # Scenes where this model performs best
                        'worst_scenes': []  # Scenes where this model performs worst
                    }
                
                model_performance[model_id]['gi_scores'].append(metrics['generalization_index'])
                model_performance[model_id]['coverage_ratios'].append(metrics['coverage_ratio'])
                model_performance[model_id]['success_ratios'].append(metrics['success_ratio'])
                model_performance[model_id]['hypervolumes'].append(metrics['hypervolume'])
                
                # Check if this model is the best for the scene
                if analysis['scene_comparison']['best_model'] == model_id:
                    model_performance[model_id]['best_scenes'].append(test_scenario)
                
                # Check if this model is the worst for the scene
                if analysis['scene_comparison']['worst_model'] == model_id:
                    model_performance[model_id]['worst_scenes'].append(test_scenario)
        
        # Compute overall metrics per model
        overall_results = {}
        
        for model_id, perf in model_performance.items():
            if perf['gi_scores']:
                overall_results[model_id] = {
                    'average_gi': np.mean(perf['gi_scores']),
                    'gi_std': np.std(perf['gi_scores']),
                    'min_gi': np.min(perf['gi_scores']),
                    'max_gi': np.max(perf['gi_scores']),
                    'average_coverage': np.mean(perf['coverage_ratios']),
                    'average_success': np.mean(perf['success_ratios']),
                    'average_hypervolume': np.mean(perf['hypervolumes']),
                    'best_scenes_count': len(perf['best_scenes']),
                    'worst_scenes_count': len(perf['worst_scenes']),
                    'best_scenes': perf['best_scenes'],
                    'worst_scenes': perf['worst_scenes']
                }
        
        # Ranking
        models = list(overall_results.keys())
        average_gis = [overall_results[model]['average_gi'] for model in models]
        
        if average_gis:
            ranked_indices = np.argsort(average_gis)[::-1]  # Sort high to low
            
            print("\nOverall model generalization ranking:")
            for rank, idx in enumerate(ranked_indices, 1):
                model_id = models[idx]
                gi = average_gis[idx]
                best_count = overall_results[model_id]['best_scenes_count']
                worst_count = overall_results[model_id]['worst_scenes_count']
                
                print(f"  Rank {rank}: Model {model_id} (avg GI={gi:.4f}, "
                        f"best scenes={best_count}, worst scenes={worst_count})")
        
        return overall_results


class ResultExporter:
    """Result exporter."""
    
    @staticmethod
    def export_generalization_analysis(all_scene_analysis, overall_results, output_dir):
        """Export analysis results."""
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(output_dir, f"generalization_analysis_{timestamp}.xlsx")
        
        wb = Workbook()
        
        # Sheet1: Per-scene model comparison
        ws_scenes = wb.active
        ws_scenes.title = "Scene model comparison"
        
        headers_scenes = [
            'Test scenario', 'Total solutions', 'Pareto solutions', 'Pareto ratio',
            'Best model', 'Best GI', 'Worst model', 'Worst GI', 'Performance gap',
            'Average GI', 'GI std', 'Dominance ratio'
        ]
        
        for col, header in enumerate(headers_scenes, 1):
            ws_scenes.cell(row=1, column=col, value=header)
        
        row = 2
        for test_scenario, analysis in all_scene_analysis.items():
            if analysis is None:
                continue
                
            scene_comparison = analysis['scene_comparison']
            model_metrics = analysis['model_metrics']
            
            best_model = scene_comparison['best_model']
            worst_model = scene_comparison['worst_model']
            
            best_gi = model_metrics[best_model]['generalization_index'] if best_model in model_metrics else 0
            worst_gi = model_metrics[worst_model]['generalization_index'] if worst_model in model_metrics else 0
            
            ws_scenes.cell(row=row, column=1, value=test_scenario)
            ws_scenes.cell(row=row, column=2, value=analysis['total_solutions'])
            ws_scenes.cell(row=row, column=3, value=analysis['total_pareto_solutions'])
            ws_scenes.cell(row=row, column=4, value=analysis['pareto_ratio'])
            ws_scenes.cell(row=row, column=5, value=best_model)
            ws_scenes.cell(row=row, column=6, value=best_gi)
            ws_scenes.cell(row=row, column=7, value=worst_model)
            ws_scenes.cell(row=row, column=8, value=worst_gi)
            ws_scenes.cell(row=row, column=9, value=scene_comparison['performance_gap'])
            ws_scenes.cell(row=row, column=10, value=scene_comparison['average_gi'])
            ws_scenes.cell(row=row, column=11, value=scene_comparison['gi_std'])
            ws_scenes.cell(row=row, column=12, value=scene_comparison['dominance_ratio'])
            
            row += 1
        
        # Sheet2: Per-model metrics by scene
        ws_models = wb.create_sheet("Model metrics by scene")
        
        headers_models = [
            'Test scenario', 'Training model', 'Solution count', 'Pareto contributions',
            'Coverage', 'Success', 'Hypervolume (HV)', 'Spacing', 'Convergence',
            'Generalization index (GI)', 'Generalization level'
        ]
        
        for col, header in enumerate(headers_models, 1):
            ws_models.cell(row=1, column=col, value=header)
        
        row = 2
        for test_scenario, analysis in all_scene_analysis.items():
            if analysis is None:
                continue
                
            model_metrics = analysis['model_metrics']
            
            for model_id, metrics in model_metrics.items():
                ws_models.cell(row=row, column=1, value=test_scenario)
                ws_models.cell(row=row, column=2, value=model_id)
                ws_models.cell(row=row, column=3, value=metrics['solution_count'])
                ws_models.cell(row=row, column=4, value=metrics['pareto_count'])
                ws_models.cell(row=row, column=5, value=metrics['coverage_ratio'])
                ws_models.cell(row=row, column=6, value=metrics['success_ratio'])
                ws_models.cell(row=row, column=7, value=metrics['hypervolume'])
                ws_models.cell(row=row, column=8, value=metrics['spacing'])
                ws_models.cell(row=row, column=9, value=metrics['convergence'])
                ws_models.cell(row=row, column=10, value=metrics['generalization_index'])
                ws_models.cell(row=row, column=11, value=metrics['generalization_level'])
                
                row += 1
        
        # Sheet3: Group comparison analysis
        ws_groups = wb.create_sheet("Group comparison")
        
        headers_groups = [
            'Test scenario', 'Test group', 'Same-group avg GI', 'Same-group GI std',
            'Cross-group avg GI', 'Cross-group GI std', 'Generalization advantage (same - cross)'
        ]
        
        for col, header in enumerate(headers_groups, 1):
            ws_groups.cell(row=1, column=col, value=header)
        
        row = 2
        for test_scenario, analysis in all_scene_analysis.items():
            if analysis is None or 'group_analysis' not in analysis:
                continue
                
            group_analysis = analysis['group_analysis']
            
            ws_groups.cell(row=row, column=1, value=test_scenario)
            ws_groups.cell(row=row, column=2, value=group_analysis['test_scenario_group'])
            ws_groups.cell(row=row, column=3, value=group_analysis['same_group_avg'])
            ws_groups.cell(row=row, column=4, value=group_analysis['same_group_std'])
            ws_groups.cell(row=row, column=5, value=group_analysis['different_group_avg'])
            ws_groups.cell(row=row, column=6, value=group_analysis['different_group_std'])
            ws_groups.cell(row=row, column=7, value=group_analysis['generalization_advantage'])
            
            row += 1
        
        # Sheet4: Overall model performance
        if overall_results:
            ws_overall = wb.create_sheet("Overall model performance")
            
            headers_overall = [
                'Training model', 'Average GI', 'GI std', 'Min GI', 'Max GI',
                'Average coverage', 'Average success', 'Average HV',
                'Best scenes count', 'Worst scenes count', 'Best scenes', 'Worst scenes'
            ]
            
            for col, header in enumerate(headers_overall, 1):
                ws_overall.cell(row=1, column=col, value=header)
            
            row = 2
            for model_id, metrics in overall_results.items():
                ws_overall.cell(row=row, column=1, value=model_id)
                ws_overall.cell(row=row, column=2, value=metrics['average_gi'])
                ws_overall.cell(row=row, column=3, value=metrics['gi_std'])
                ws_overall.cell(row=row, column=4, value=metrics['min_gi'])
                ws_overall.cell(row=row, column=5, value=metrics['max_gi'])
                ws_overall.cell(row=row, column=6, value=metrics['average_coverage'])
                ws_overall.cell(row=row, column=7, value=metrics['average_success'])
                ws_overall.cell(row=row, column=8, value=metrics['average_hypervolume'])
                ws_overall.cell(row=row, column=9, value=metrics['best_scenes_count'])
                ws_overall.cell(row=row, column=10, value=metrics['worst_scenes_count'])
                ws_overall.cell(row=row, column=11, value=str(metrics['best_scenes']))
                ws_overall.cell(row=row, column=12, value=str(metrics['worst_scenes']))
                
                row += 1
        
        # Save workbook
        wb.save(filename)
        print(f"Generalization analysis saved to: {filename}")
        
        return filename


# ============================================
# 3. Visualization generation
# ============================================

class VisualizationGenerator:
    """Visualization generator."""
    
    @staticmethod
    def plot_generalization_bar_chart(all_scene_analysis, output_dir):
        """Generate a generalization bar chart across test scenarios."""
        print("\nGenerating generalization bar chart...")
        
        # Define scene order and groups
        scenes = ['L9-1', 'L9-2', 'L9-3', 'L9-4', 'L9-5', 'L9-6', 'L9-7', 'L9-8', 'L9-9']
        scene_groups = ['B', 'A', 'B', 'C', 'A', 'A', 'C', 'B', 'B']
        
        # Training model mapping
        model_names = {
            'L9-6': 'Model_A (S6/Group A)',
            'L9-8': 'Model_B (S8/Group B)',
            'L9-4': 'Model_C (S4/Group C)'
        }
        
        # Collect GI values per model
        model_gi_values = {
            'L9-6': [],  # Model_A
            'L9-8': [],  # Model_B
            'L9-4': []   # Model_C
        }
        
        # Collect data from analysis results
        for scene in scenes:
            if scene in all_scene_analysis and all_scene_analysis[scene]:
                model_metrics = all_scene_analysis[scene]['model_metrics']
                
                for model_id in ['L9-6', 'L9-8', 'L9-4']:
                    if model_id in model_metrics:
                        gi = model_metrics[model_id]['generalization_index']
                        model_gi_values[model_id].append(gi)
                    else:
                        model_gi_values[model_id].append(0)
            else:
                # Use 0 as a placeholder if no analysis result
                for model_id in ['L9-6', 'L9-8', 'L9-4']:
                    model_gi_values[model_id].append(0)
        
        # Create figure
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 10), 
                                        gridspec_kw={'height_ratios': [3, 1], 'hspace': 0.08})
        
        x = np.arange(len(scenes))
        width = 0.25
        
        # Colors
        colors = {'L9-6': '#1f77b4', 'L9-8': '#ff7f0e', 'L9-4': '#2ca02c'}
        
        # Draw bars
        bars = []
        for i, (model_id, values) in enumerate(model_gi_values.items()):
            offset = (i - 1) * width
            bar = ax1.bar(x + offset, values, width, 
                          label=model_names[model_id],
                          color=colors[model_id], 
                          edgecolor='black', 
                          linewidth=0.5, 
                          alpha=0.8)
            bars.append(bar)
            
            # Annotate values above bars
            for j, v in enumerate(values):
                if v > 0:
                    ax1.text(x[j] + offset, v + 0.02, f'{v:.2f}', 
                            ha='center', va='bottom', fontsize=8, rotation=0)
        
        # Main chart styling
        ax1.set_ylabel('Generalization Index (GI)', fontsize=12, fontweight='bold')
        ax1.set_xticks(x)
        ax1.set_xticklabels(scenes, fontsize=11)
        ax1.set_ylim(0, 1.2)
        ax1.grid(True, axis='y', linestyle='--', alpha=0.3)
        ax1.legend(loc='upper right', fontsize=10)
        ax1.set_title('Generalization comparison across test scenes', fontsize=14, fontweight='bold', pad=15)
        
        # Add scene group labels
        for i, group in enumerate(scene_groups):
            ax1.text(x[i], -0.08, f'(Group {group})', ha='center', va='top', fontsize=9, color='gray')
        
        # Secondary plot: best model per scene
        best_models = []
        for scene in scenes:
            if scene in all_scene_analysis and all_scene_analysis[scene]:
                best = all_scene_analysis[scene]['scene_comparison']['best_model']
                if best == 'L9-6':
                    best_models.append(0)  # Model_A
                elif best == 'L9-8':
                    best_models.append(1)  # Model_B
                elif best == 'L9-4':
                    best_models.append(2)  # Model_C
                else:
                    best_models.append(-1)
            else:
                best_models.append(-1)
        
        # Mark the best model in the secondary plot
        for i, best in enumerate(best_models):
            if best == 0:
                ax2.bar(x[i], 1, width, color=colors['L9-6'], alpha=0.6, edgecolor='black', linewidth=0.5)
                ax2.text(x[i], 0.5, 'Model_A', ha='center', va='center', fontsize=8)
            elif best == 1:
                ax2.bar(x[i], 1, width, color=colors['L9-8'], alpha=0.6, edgecolor='black', linewidth=0.5)
                ax2.text(x[i], 0.5, 'Model_B', ha='center', va='center', fontsize=8)
            elif best == 2:
                ax2.bar(x[i], 1, width, color=colors['L9-4'], alpha=0.6, edgecolor='black', linewidth=0.5)
                ax2.text(x[i], 0.5, 'Model_C', ha='center', va='center', fontsize=8)
            else:
                ax2.bar(x[i], 1, width, color='gray', alpha=0.3, edgecolor='black', linewidth=0.5)
                ax2.text(x[i], 0.5, 'N/A', ha='center', va='center', fontsize=8)
        
        ax2.set_ylabel('Best model', fontsize=11)
        ax2.set_xticks(x)
        ax2.set_xticklabels(scenes, fontsize=11)
        ax2.set_ylim(0, 1.2)
        ax2.set_yticks([])
        
        # Add average GI annotation
        avg_gi = {}
        for model_id in ['L9-6', 'L9-8', 'L9-4']:
            values = [v for v in model_gi_values[model_id] if v > 0]
            avg_gi[model_id] = np.mean(values) if values else 0
        
        stats_text = f"Avg GI: Model_A={avg_gi['L9-6']:.3f}, Model_B={avg_gi['L9-8']:.3f}, Model_C={avg_gi['L9-4']:.3f}"
        ax1.text(0.02, 0.95, stats_text, transform=ax1.transAxes, fontsize=10,
                 bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
        
        # Save images
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(output_dir, f"generalization_bar_chart_{timestamp}.png")
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        print(f"Bar chart saved to: {filename}")
        
        # Also save as PDF (vector, suitable for papers)
        pdf_filename = os.path.join(output_dir, f"generalization_bar_chart_{timestamp}.pdf")
        plt.savefig(pdf_filename, bbox_inches='tight')
        print(f"Vector figure saved to: {pdf_filename}")
        
        plt.close()
        return filename

    @staticmethod
    def plot_generalization_heatmap(all_scene_analysis, output_dir):
        """Generate a generalization heatmap across scenes and models."""
        print("\nGenerating generalization heatmap...")
        
        scenes = ['L9-1', 'L9-2', 'L9-3', 'L9-4', 'L9-5', 'L9-6', 'L9-7', 'L9-8', 'L9-9']
        models = ['L9-6', 'L9-8', 'L9-4']
        model_labels = ['Model_A\n(S6/Group A)', 'Model_B\n(S8/Group B)', 'Model_C\n(S4/Group C)']
        
        # Build data matrix
        data = np.zeros((len(models), len(scenes)))
        
        for i, model_id in enumerate(models):
            for j, scene in enumerate(scenes):
                if scene in all_scene_analysis and all_scene_analysis[scene]:
                    model_metrics = all_scene_analysis[scene]['model_metrics']
                    if model_id in model_metrics:
                        data[i, j] = model_metrics[model_id]['generalization_index']
        
        # Create figure
        fig, ax = plt.subplots(figsize=(14, 6))
        
        # Draw heatmap
        im = ax.imshow(data, cmap='YlOrRd', aspect='auto', vmin=0, vmax=1.1)
        
        # Axis labels
        ax.set_xticks(np.arange(len(scenes)))
        ax.set_xticklabels(scenes, fontsize=11)
        ax.set_yticks(np.arange(len(models)))
        ax.set_yticklabels(model_labels, fontsize=11)
        
        # Annotate each cell
        for i in range(len(models)):
            for j in range(len(scenes)):
                if data[i, j] > 0:
                    text = ax.text(j, i, f'{data[i, j]:.2f}',
                                  ha='center', va='center', color='black', fontsize=9)
        
        # Add colorbar
        cbar = plt.colorbar(im, ax=ax, shrink=0.8)
        cbar.set_label('Generalization Index (GI)', fontsize=11)
        
        # Title and labels
        ax.set_title('Generalization heatmap across test scenes', fontsize=14, fontweight='bold', pad=15)
        ax.set_xlabel('Test scenes', fontsize=12)
        ax.set_ylabel('Training models', fontsize=12)
        
        # Add scene group labels
        scene_groups = ['B', 'A', 'B', 'C', 'A', 'A', 'C', 'B', 'B']
        for j, group in enumerate(scene_groups):
            ax.text(j, -0.2, f'(Group {group})', ha='center', va='center', fontsize=9, color='gray')
        
        plt.tight_layout()
        
        # Save images
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(output_dir, f"generalization_heatmap_{timestamp}.png")
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        print(f"Heatmap saved to: {filename}")
        
        pdf_filename = os.path.join(output_dir, f"generalization_heatmap_{timestamp}.pdf")
        plt.savefig(pdf_filename, bbox_inches='tight')
        print(f"Vector figure saved to: {pdf_filename}")
        
        plt.close()
        return filename

    @staticmethod
    def plot_model_performance_radar(overall_results, output_dir):
        """Generate a radar chart for overall model performance."""
        print("\nGenerating overall model performance radar chart...")
        
        if not overall_results:
            print("No overall results available; skip radar chart")
            return None
        
        models = list(overall_results.keys())
        model_labels = [f'Model_{m[-1]}\n({m})' for m in models]
        
        # Extract metrics
        metrics = ['Avg GI', 'Avg coverage', 'Avg success', 'Avg HV']
        values = []
        
        for model in models:
            model_metrics = overall_results[model]
            values.append([
                model_metrics['average_gi'],
                model_metrics['average_coverage'],
                model_metrics['average_success'],
                model_metrics['average_hypervolume']
            ])
        
        # Normalize
        values = np.array(values)
        normalized_values = np.zeros_like(values)
        
        for i in range(values.shape[1]):
            col = values[:, i]
            if np.max(col) > np.min(col):
                normalized_values[:, i] = (col - np.min(col)) / (np.max(col) - np.min(col))
            else:
                normalized_values[:, i] = 0.5
        
        # Radar chart angles
        angles = np.linspace(0, 2 * np.pi, len(metrics), endpoint=False).tolist()
        angles += angles[:1]  # Close the loop
        
        # Create figure
        fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(projection='polar'))
        
        colors = ['#1f77b4', '#ff7f0e', '#2ca02c']
        
        for i, model in enumerate(models):
            model_values = normalized_values[i].tolist()
            model_values += model_values[:1]  # Close the loop
            
            ax.plot(angles, model_values, 'o-', linewidth=2, color=colors[i], label=model_labels[i])
            ax.fill(angles, model_values, alpha=0.1, color=colors[i])
        
        # Angle labels
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(metrics, fontsize=11)
        
        # Y-axis range
        ax.set_ylim(0, 1)
        ax.set_rlabel_position(30)
        
        # Legend
        ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.0), fontsize=10)
        
        ax.set_title('Overall model performance (normalized)', fontsize=14, fontweight='bold', pad=20)
        
        plt.tight_layout()
        
        # Save images
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(output_dir, f"model_radar_{timestamp}.png")
        plt.savefig(filename, dpi=300, bbox_inches='tight')
        print(f"Radar chart saved to: {filename}")
        
        pdf_filename = os.path.join(output_dir, f"model_radar_{timestamp}.pdf")
        plt.savefig(pdf_filename, bbox_inches='tight')
        print(f"Vector figure saved to: {pdf_filename}")
        
        plt.close()
        return filename


# ============================================
# 4. Core experiment functions
# ============================================

def find_or_train_model(scenario_id, algorithm='PG_HRL'):
    """Find or train the model for a given scenario."""
    scenario = None
    for s in orthogonal_scenarios:
        if s['scenario_id'] == scenario_id:
            scenario = s
            break
    
    if not scenario:
        print(f"Error: scenario not found {scenario_id}")
        return None    
    
    # Build model path
    model_address = (
        f"{sys.path[0]}/sequencing_models/"
        f"{algorithm}_{scenario_id}.pt"
    )
    
    if os.path.exists(model_address):
        print(f"Found trained model: {model_address}")
        return model_address
    else:
        print(f"Model not found, start training scenario {scenario_id}...")
        return None

def run_model_on_scene(train_scenario_id, test_scenario_id, model_address):
    """Run a model on one test scenario."""
    
    # Get test scenario
    test_scenario = None
    for s in orthogonal_scenarios:
        if s['scenario_id'] == test_scenario_id:
            test_scenario = s
            break
    
    if not test_scenario:
        print(f"Error: test scenario not found {test_scenario_id}")
        return None
    
    test_params = test_scenario['parameters']
   
    # Initialize multi-objective manager
    mo_manager = MultiObjectiveManager()
    
    print(f"  Model {train_scenario_id} -> Scene {test_scenario_id}")
    
    # Repeated tests
    for rep in range(EXPERIMENT_CONFIG['test_repetitions']):
        seed = np.random.randint(EXPERIMENT_CONFIG['random_seed_range'])
        # Generate perturbed preference vectors
        perturbed_preferences = torch.tensor([np.random.dirichlet([1, 1, 1]) for _ in range(EXPERIMENT_CONFIG['preference_perturbations'])],dtype=torch.float32) 
        
        # Run tests for each preference vector
        for pref_idx, perturbed_pref in enumerate(perturbed_preferences):
            env = simpy.Environment()
            spf = ShopFloor(env, EXPERIMENT_CONFIG['simulation_span'],test_scenario['parameters'],
                            brain_machine, seed=seed,preference_vector=perturbed_pref, 
                            address=model_address)           
            # Run simulation
            spf.simulation()
            # Collect results
            run_id = f"train_{train_scenario_id}_test_{test_scenario_id}_rep_{rep}_pref_{pref_idx}"
            mo_manager.add_experiment_data(ALGORITHM, run_id, spf.job_objectives_records  )
    
    # Collect all solutions
    all_solutions = []
    for rule_data in mo_manager.all_experiment_data.values():
        for run_data in rule_data.values():
            for job_data in run_data:
                if 'objectives' in job_data:
                    all_solutions.append(job_data['objectives'])
    
    return {
        'train_scenario': train_scenario_id,
        'test_scenario': test_scenario_id,
        'solution_count': len(all_solutions),
        'raw_data': mo_manager.all_experiment_data
    }

# ============================================
# 5. Main experiment flow
# ============================================

def main():   
    """Main experiment function."""
    print("=" * 60)
    print("DRL generalization experiment (fixed scene, compare models)")
    print(f"Algorithm: {ALGORITHM}")
    print(f"Training models: {list(REPRESENTATIVE_MODELS.values())}")
    print("Test scenes: all 9 orthogonal scenarios")
    print("=" * 60)
    
    # Ensure output directory exists
    output_dir = os.path.join(sys.path[0], "test_result\\Robust")
    os.makedirs(output_dir, exist_ok=True)
    
    # Store all results
    all_model_results = {}  # Test results grouped by model
    all_scene_analysis = {}  # Analysis results grouped by scene
    
    # ============================================
    # Stage 1: Train all models
    # ============================================
    print(f"\n{'='*50}")
    print("Stage 1: Train all models")
    print(f"{'='*50}")
    
    for group_name, train_scenario_id in REPRESENTATIVE_MODELS.items():
        print(f"\nProcessing training model: {train_scenario_id} (Group: {group_name})")
        
        # Find or train model
        model_address = find_or_train_model(train_scenario_id, ALGORITHM)
        if not model_address:
            print(f"Warning: model unavailable {train_scenario_id}, skip")
            continue
    
    # ============================================
    # Stage 2: Collect test data (outer loop: scenes)
    # ============================================
    print(f"\n{'='*50}")
    print("Stage 2: Collect test data")
    print(f"{'='*50}")
    
    # Collect all models on all scenes
    all_model_results = {}  # {train_model_id: {test_scenario_id: results}}
    
    # For each training model
    for train_scenario_id in REPRESENTATIVE_MODELS.values():
        print(f"\nCollecting data for model {train_scenario_id}...")
        model_address = find_or_train_model(train_scenario_id, ALGORITHM)
        if not model_address:
            continue
        
        all_model_results[train_scenario_id] = {}
        
        # Run the model on all 9 scenes
        for test_scenario in orthogonal_scenarios:
            test_scenario_id = test_scenario['scenario_id']
            
            result = run_model_on_scene(train_scenario_id, test_scenario_id, model_address)
            if result:
                all_model_results[train_scenario_id][test_scenario_id] = result
                print(f"  Scene {test_scenario_id}: {result['solution_count']} solutions")
    
    # ============================================
    # Stage 3: Analyze by scene (outer loop: scenes)
    # ============================================
    print(f"\n{'='*50}")
    print("Stage 3: Analyze model adaptability per scene")
    print(f"{'='*50}")
    
    # Analyze each test scene
    for test_scenario in orthogonal_scenarios:
        test_scenario_id = test_scenario['scenario_id']
        
        print(f"\nAnalyzing scene: {test_scenario_id}")
        
        # Analyze generalization across models for this scene
        scene_analysis = SceneGeneralizationAnalyzer.analyze_scene_generalization(
            test_scenario_id, 
            all_model_results
        )
        
        if scene_analysis:
            all_scene_analysis[test_scenario_id] = scene_analysis
    
    # ============================================
    # Stage 4: Cross-scene synthesis
    # ============================================
    print(f"\n{'='*60}")
    print("Stage 4: Cross-scene synthesis")
    print(f"{'='*60}")
    
    overall_results = CrossSceneAnalyzer.analyze_overall_generalization(all_scene_analysis)
    
    # ============================================
    # Stage 5: Generate visualizations
    # ============================================
    print(f"\n{'='*60}")
    print("Stage 5: Generate visualizations")
    print(f"{'='*60}")
    
    try:
        # Bar chart
        if all_scene_analysis:
            bar_chart_file = VisualizationGenerator.plot_generalization_bar_chart(all_scene_analysis, output_dir)
            
            # Heatmap
            heatmap_file = VisualizationGenerator.plot_generalization_heatmap(all_scene_analysis, output_dir)
            
            # Radar chart
            if overall_results:
                radar_file = VisualizationGenerator.plot_model_performance_radar(overall_results, output_dir)
            
            print("\nAll visualizations generated")
    except Exception as e:
        print(f"Error while generating visualizations: {e}")
        import traceback
        traceback.print_exc()
    
    # ============================================
    # Stage 6: Export results
    # ============================================
    print(f"\n{'='*60}")
    print("Stage 6: Export analysis results")
    print(f"{'='*60}")
    
    try:
        # Export analysis results
        if all_scene_analysis:
            analysis_file = ResultExporter.export_generalization_analysis(
                all_scene_analysis, overall_results, output_dir
            )
            print(f"Generalization analysis saved to: {analysis_file}")
        
        # Save raw test data
        raw_data_file = os.path.join(output_dir, f"raw_test_data_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pkl")
        with open(raw_data_file, 'wb') as f:
            pickle.dump(all_model_results, f)
        print(f"Raw test data saved to: {raw_data_file}")
        
        print(f"\nAll results saved to directory: {output_dir}")
        
    except Exception as e:
        print(f"Error while exporting results: {e}")
        import traceback
        traceback.print_exc()
    
    # ============================================
    # Stage 7: Experiment summary
    # ============================================
    print(f"\n{'='*60}")
    print("Experiment summary")
    print(f"{'='*60}")
    
    # Best model per scene
    print("\nBest-adapting model per scene:")
    for test_scenario, analysis in all_scene_analysis.items():
        if analysis:
            best_model = analysis['scene_comparison']['best_model']
            if best_model:
                best_gi = analysis['model_metrics'][best_model]['generalization_index']
                print(f"  Scene {test_scenario}: {best_model} (GI={best_gi:.4f})")
    
    # Overall model performance
    if overall_results:
        print("\nOverall model generalization:")
        for model_id, metrics in overall_results.items():
            print(f"  Model {model_id}: avg GI={metrics['average_gi']:.4f}, "
                f"best scenes={metrics['best_scenes_count']}, "
                f"worst scenes={metrics['worst_scenes_count']}")
    
    return {
        'model_results': all_model_results,
        'scene_analysis': all_scene_analysis,
        'overall_results': overall_results
    }

# ============================================
# 6. Program entry
# ============================================

if __name__ == "__main__":
    # Set random seed for reproducibility
    np.random.seed(42)
    
    try:
        results = main()
        print("\nGeneralization experiment (new approach) completed!")
    except KeyboardInterrupt:
        print("\nExperiment interrupted by user")
    except Exception as e:
        print(f"\nError during experiment: {e}")
        import traceback
        traceback.print_exc()